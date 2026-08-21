from __future__ import annotations

from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook
from PIL import Image

import scripts.render_portfolio_screenshots as renderer
from scripts.render_portfolio_screenshots import (
    PortfolioRenderError,
    render_portfolio_screenshots,
)


EXPECTED_ASSETS = (
    "01-summary.png",
    "02-valid-records.png",
    "03-invalid-records.png",
    "04-duplicates.png",
)


def save_portfolio_workbook(path: Path) -> Path:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Total Records", 3])
    summary.append(["Valid Records", 1])
    summary.append(["Invalid Records", 1])
    summary.append(["Duplicate Records", 1])
    summary.append(["Total Paid Amount (USD)", "10.00"])
    summary.append(["Invalid and duplicate record counts may overlap.", None])

    headers = [
        "order_id",
        "customer_name",
        "email",
        "order_date",
        "amount",
        "status",
        "note",
        "source_file",
        "source_sheet",
        "source_row",
    ]
    valid = workbook.create_sheet("Valid Records")
    valid.append(headers)
    valid.append(
        [
            "001",
            "Ada Example",
            "ada@example.com",
            "2026-08-01",
            "10.00",
            "paid",
            None,
            "orders.csv",
            None,
            2,
        ]
    )

    review_headers = [*headers, "validation_errors"]
    invalid = workbook.create_sheet("Invalid Records")
    invalid.append(review_headers)
    invalid.append(
        [
            "BAD",
            "Cy Example",
            "invalid",
            "2026-08-02",
            "2.00",
            "paid",
            None,
            "orders.csv",
            None,
            3,
            "email [invalid_email]: email format is invalid",
        ]
    )
    duplicates = workbook.create_sheet("Duplicates")
    duplicates.append(review_headers)
    duplicates.append(
        [
            "DUP",
            "Ben Example",
            "ben@example.com",
            "2026-08-03",
            "3.00",
            "pending",
            None,
            "orders.xlsx",
            "Orders",
            2,
            None,
        ]
    )
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def rendered_assets(tmp_path: Path) -> tuple[Path, Path]:
    workbook = save_portfolio_workbook(tmp_path / "sales-report.xlsx")
    output = tmp_path / "screenshots"
    render_portfolio_screenshots(workbook, output)
    return workbook, output


def test_validate_only_accepts_exact_deterministic_assets(
    rendered_assets: tuple[Path, Path],
) -> None:
    workbook, output = rendered_assets

    render_portfolio_screenshots(workbook, output, validate_only=True)


def test_validate_only_rejects_white_pngs(
    rendered_assets: tuple[Path, Path],
) -> None:
    workbook, output = rendered_assets
    for filename in EXPECTED_ASSETS:
        path = output / filename
        with Image.open(path) as image:
            Image.new("RGB", image.size, "white").save(path, format="PNG")

    with pytest.raises(PortfolioRenderError, match="content differs"):
        render_portfolio_screenshots(workbook, output, validate_only=True)


def test_validate_only_rejects_one_changed_pixel(
    rendered_assets: tuple[Path, Path],
) -> None:
    workbook, output = rendered_assets
    path = output / "01-summary.png"
    with Image.open(path) as image:
        changed = image.copy()
    changed.putpixel((0, 0), (0, 0, 0))
    changed.save(path, format="PNG", optimize=True)

    with pytest.raises(PortfolioRenderError, match="content differs"):
        render_portfolio_screenshots(workbook, output, validate_only=True)


def test_validate_only_rejects_asset_under_wrong_filename(
    rendered_assets: tuple[Path, Path],
) -> None:
    workbook, output = rendered_assets
    valid = (output / "02-valid-records.png").read_bytes()
    invalid = (output / "03-invalid-records.png").read_bytes()
    (output / "02-valid-records.png").write_bytes(invalid)
    (output / "03-invalid-records.png").write_bytes(valid)

    with pytest.raises(PortfolioRenderError):
        render_portfolio_screenshots(workbook, output, validate_only=True)


def test_validate_only_rejects_valid_but_incorrect_png(
    rendered_assets: tuple[Path, Path],
) -> None:
    workbook, output = rendered_assets
    path = output / "04-duplicates.png"
    with Image.open(path) as image:
        Image.new("RGB", image.size, "gray").save(path, format="PNG")

    with pytest.raises(PortfolioRenderError, match="content differs"):
        render_portfolio_screenshots(workbook, output, validate_only=True)


def test_validate_only_rejects_wrong_dimensions(
    rendered_assets: tuple[Path, Path],
) -> None:
    workbook, output = rendered_assets
    Image.new("RGB", (10, 10), "white").save(
        output / "01-summary.png", format="PNG"
    )

    with pytest.raises(PortfolioRenderError, match="dimensions"):
        render_portfolio_screenshots(workbook, output, validate_only=True)


def test_validate_only_rejects_corrupted_asset(
    rendered_assets: tuple[Path, Path],
) -> None:
    workbook, output = rendered_assets
    (output / "01-summary.png").write_bytes(b"not a PNG")

    with pytest.raises(PortfolioRenderError, match="cannot be read"):
        render_portfolio_screenshots(workbook, output, validate_only=True)


def test_validate_only_rejects_missing_asset(
    rendered_assets: tuple[Path, Path],
) -> None:
    workbook, output = rendered_assets
    (output / "01-summary.png").unlink()

    with pytest.raises(PortfolioRenderError, match="missing or empty"):
        render_portfolio_screenshots(workbook, output, validate_only=True)


def test_rendering_is_byte_deterministic_and_does_not_modify_workbook(
    tmp_path: Path,
) -> None:
    workbook = save_portfolio_workbook(tmp_path / "sales-report.xlsx")
    before = sha256(workbook.read_bytes()).digest()
    first = tmp_path / "first"
    second = tmp_path / "second"

    render_portfolio_screenshots(workbook, first)
    render_portfolio_screenshots(workbook, second)

    assert all(
        (first / filename).read_bytes() == (second / filename).read_bytes()
        for filename in EXPECTED_ASSETS
    )
    assert sha256(workbook.read_bytes()).digest() == before


def test_reports_clear_error_when_dejavu_fonts_are_unavailable(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = save_portfolio_workbook(tmp_path / "sales-report.xlsx")
    monkeypatch.setattr(renderer, "FONT_REGULAR_CANDIDATES", ("missing.ttf",))
    monkeypatch.setattr(renderer, "FONT_BOLD_CANDIDATES", ("missing-bold.ttf",))

    with pytest.raises(
        PortfolioRenderError,
        match="required DejaVu Sans .* font is unavailable",
    ):
        render_portfolio_screenshots(workbook, tmp_path / "screenshots")
