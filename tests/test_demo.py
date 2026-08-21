from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from src.exporter import export_report
from src.processor import (
    discover_supported_files,
    load_supported_files,
    process_records,
)
from src.summary import calculate_summary


DEMO_INPUT_DIR = Path("data/demo/input")


def test_fictitious_demo_matches_documented_result(tmp_path: Path) -> None:
    sources = discover_supported_files(DEMO_INPUT_DIR)
    assert [path.name for path in sources] == ["orders.csv", "orders.xlsx"]

    result = process_records(load_supported_files(sources))
    summary = calculate_summary(result)

    assert summary.total_records == 11
    assert summary.valid_records == 5
    assert summary.invalid_records == 4
    assert summary.duplicate_records == 4
    assert summary.total_paid_amount == Decimal("34.46")
    assert [record.record["order_id"] for record in result.valid_records] == [
        "001",
        "XLSX-PAID",
        "0007",
        "CANCEL-0",
        "ROUND-PAID",
    ]
    assert [record.record["order_id"] for record in result.invalid_records] == [
        "BAD-EMAIL",
        "BOTH-200",
        "BOTH-200",
        "FORMULA-1",
    ]
    assert [record.record["order_id"] for record in result.duplicate_records] == [
        "DUP-100",
        "BOTH-200",
        "DUP-100",
        "BOTH-200",
    ]

    output_path = tmp_path / "sales_report.xlsx"
    export_report(result, summary, output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        assert workbook.sheetnames == [
            "Summary",
            "Valid Records",
            "Invalid Records",
            "Duplicates",
        ]
        assert [workbook["Summary"].cell(row, 2).value for row in range(2, 7)] == [
            11,
            5,
            4,
            4,
            "34.46",
        ]

        valid = workbook["Valid Records"]
        valid_headers = [cell.value for cell in valid[1]]
        valid_ids = [
            valid.cell(row, valid_headers.index("order_id") + 1).value
            for row in range(2, valid.max_row + 1)
        ]
        assert "0007" in valid_ids
        csv_row = valid_ids.index("001") + 2
        assert valid.cell(csv_row, valid_headers.index("source_file") + 1).value == (
            "orders.csv"
        )
        assert valid.cell(csv_row, valid_headers.index("source_sheet") + 1).value is None
        assert valid.cell(csv_row, valid_headers.index("source_row") + 1).value == 2
        csv_note = valid.cell(csv_row, valid_headers.index("note") + 1)
        assert csv_note.value == "'=Quarterly target"
        assert csv_note.data_type != "f"

        invalid = workbook["Invalid Records"]
        invalid_headers = [cell.value for cell in invalid[1]]
        invalid_ids = [
            invalid.cell(row, invalid_headers.index("order_id") + 1).value
            for row in range(2, invalid.max_row + 1)
        ]
        formula_row = invalid_ids.index("FORMULA-1") + 2
        formula_note = invalid.cell(formula_row, invalid_headers.index("note") + 1)
        assert formula_note.value == "'=1+1"
        assert formula_note.data_type != "f"
        assert "formula" in invalid.cell(
            formula_row,
            invalid_headers.index("validation_errors") + 1,
        ).value.lower()

        for worksheet in workbook.worksheets:
            assert all(
                cell.data_type != "f"
                for row in worksheet.iter_rows()
                for cell in row
            )
    finally:
        workbook.close()
