from __future__ import annotations

import logging
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook as open_workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from src.exporter import export_report
from src.processor import (
    NoSupportedInputFilesError,
    XlsxStructuralError,
    discover_supported_files,
    load_supported_files,
    load_xlsx_file,
    process_records,
)
from src.summary import calculate_summary
from src.validator import FormulaValue


VALID_HEADER = [
    "order_id",
    "customer_name",
    "email",
    "order_date",
    "amount",
    "status",
]
VALID_ROW = [
    "00123",
    "Julia",
    "julia@example.com",
    "2026-08-20",
    "10.00",
    "paid",
]


def save_workbook(path: Path, sheets: list[tuple[str, list[list[object]]]]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        worksheet = workbook.create_sheet(title)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def save_special_formula_workbook(path: Path) -> Path:
    """Create a real workbook containing every formula representation we support."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders"
    worksheet.append([*VALID_HEADER, "note"])
    worksheet.append([*VALID_ROW, "=SUM(1,2)"])
    worksheet.append(
        ["ARRAY", "Array", None, "2026-08-21", "2.00", "paid", None]
    )
    worksheet["G3"] = ArrayFormula("G3:G3", "=SUM(2,3)")
    worksheet.append(
        ["TABLE", "Table", None, "2026-08-22", "3.00", "paid", None]
    )
    worksheet["G4"] = DataTableFormula(
        "G4:G4",
        ca=True,
        dt2D=True,
        dtr=False,
        r1="A1",
        r2="B1",
        del1=False,
        del2=True,
    )
    workbook.save(path)
    workbook.close()
    return path


@pytest.mark.parametrize("name", ["orders.xlsx", "orders.XLSX", "orders.XlSx"])
def test_discovers_xlsx_extension_case_insensitively(tmp_path: Path, name: str) -> None:
    source = save_workbook(tmp_path / name, [("Orders", [VALID_HEADER, VALID_ROW])])

    assert discover_supported_files(tmp_path) == [source]


def test_discovers_csv_and_xlsx_in_deterministic_order(tmp_path: Path) -> None:
    (tmp_path / "b.csv").write_text("content", encoding="utf-8")
    save_workbook(tmp_path / "A.XLSX", [("Orders", [VALID_HEADER, VALID_ROW])])

    assert [path.name for path in discover_supported_files(tmp_path)] == [
        "A.XLSX",
        "b.csv",
    ]


def test_xlsx_discovery_is_non_recursive_and_ignores_symlinks(tmp_path: Path) -> None:
    external = save_workbook(
        tmp_path.parent / "external.xlsx",
        [("Orders", [VALID_HEADER, VALID_ROW])],
    )
    (tmp_path / "linked.xlsx").symlink_to(external)
    nested = tmp_path / "nested"
    nested.mkdir()
    save_workbook(nested / "nested.xlsx", [("Orders", [VALID_HEADER, VALID_ROW])])
    regular = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [VALID_HEADER, VALID_ROW])],
    )

    assert discover_supported_files(tmp_path) == [regular]


def test_only_xlsx_symlinks_means_no_supported_input(tmp_path: Path) -> None:
    external = save_workbook(
        tmp_path.parent / "external-only.xlsx",
        [("Orders", [VALID_HEADER, VALID_ROW])],
    )
    (tmp_path / "linked.xlsx").symlink_to(external)

    with pytest.raises(NoSupportedInputFilesError):
        discover_supported_files(tmp_path)


def test_loads_complete_worksheet_with_traceability_and_extra_columns(
    tmp_path: Path,
) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [
            (
                "August Sales",
                [
                    [*VALID_HEADER, "sales_channel"],
                    [*VALID_ROW, "online"],
                ],
            )
        ],
    )

    records = load_xlsx_file(source)

    assert records == [
        {
            "order_id": "00123",
            "customer_name": "Julia",
            "email": "julia@example.com",
            "order_date": "2026-08-20",
            "amount": "10.00",
            "status": "paid",
            "sales_channel": "online",
            "source_file": "orders.xlsx",
            "source_sheet": "August Sales",
            "source_row": 2,
        }
    ]


def test_loads_multiple_complete_worksheets_in_workbook_order(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [
            ("First", [VALID_HEADER, VALID_ROW]),
            (
                "Second",
                [
                    VALID_HEADER,
                    ["002", "Ada", None, date(2026, 8, 21), 20, "pending"],
                ],
            ),
        ],
    )

    records = load_xlsx_file(source)

    assert [record["source_sheet"] for record in records] == ["First", "Second"]
    assert [record["order_id"] for record in records] == ["00123", "002"]


def test_logs_and_ignores_auxiliary_worksheet(
    tmp_path: Path,
    caplog: pytest.LogCaptureFixture,
) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [
            ("Instructions", [["note"], ["Use approved data"]]),
            ("Orders", [VALID_HEADER, VALID_ROW]),
        ],
    )

    with caplog.at_level(logging.INFO):
        records = load_xlsx_file(source)

    assert len(records) == 1
    assert "Skipped auxiliary worksheet 'Instructions'" in caplog.text
    assert "as non-data" in caplog.text


def test_partial_schema_fails_complete_workbook(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [
            ("Orders", [VALID_HEADER, VALID_ROW]),
            ("Partial", [["order_id", "status"], ["002", "paid"]]),
        ],
    )

    with pytest.raises(XlsxStructuralError, match="Partial.*missing required columns"):
        load_xlsx_file(source)


@pytest.mark.parametrize(
    "sheets",
    [
        [("Empty", [])],
        [("Auxiliary", [["notes"], ["text"]])],
        [("Header Only", [VALID_HEADER])],
    ],
)
def test_workbook_without_usable_worksheet_fails(
    tmp_path: Path,
    sheets: list[tuple[str, list[list[object]]]],
) -> None:
    source = save_workbook(tmp_path / "orders.xlsx", sheets)

    with pytest.raises(XlsxStructuralError, match="no usable worksheet"):
        load_xlsx_file(source)


def test_requires_header_on_physical_row_one(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [["Sales report"], VALID_HEADER, VALID_ROW])],
    )

    with pytest.raises(XlsxStructuralError, match="header column 2 is empty"):
        load_xlsx_file(source)


@pytest.mark.parametrize(
    "replacement",
    ["Order_ID", " order_id", "order_id ", "ORDER_ID"],
)
def test_xlsx_headers_are_exact_and_case_sensitive(
    tmp_path: Path,
    replacement: str,
) -> None:
    header = [replacement if value == "order_id" else value for value in VALID_HEADER]
    source = save_workbook(tmp_path / "orders.xlsx", [("Orders", [header, VALID_ROW])])

    with pytest.raises(XlsxStructuralError, match="missing required columns: order_id"):
        load_xlsx_file(source)


@pytest.mark.parametrize(
    "missing_column",
    ["order_id", "customer_name", "order_date", "amount", "status"],
)
def test_rejects_each_missing_required_xlsx_column(
    tmp_path: Path,
    missing_column: str,
) -> None:
    header = [value for value in VALID_HEADER if value != missing_column]
    row = [
        value
        for index, value in enumerate(VALID_ROW)
        if VALID_HEADER[index] != missing_column
    ]
    source = save_workbook(tmp_path / "orders.xlsx", [("Orders", [header, row])])

    with pytest.raises(XlsxStructuralError, match=missing_column):
        load_xlsx_file(source)


def test_rejects_duplicate_nonempty_header(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [[*VALID_HEADER, "status"], [*VALID_ROW, "paid"]])],
    )

    with pytest.raises(XlsxStructuralError, match="duplicate header columns: status"):
        load_xlsx_file(source)


def test_rejects_empty_header_for_used_data_column(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [[*VALID_HEADER, None], [*VALID_ROW, "undocumented"]])],
    )

    with pytest.raises(XlsxStructuralError, match="header column 7 is empty"):
        load_xlsx_file(source)


def test_rejects_empty_internal_header_without_data(tmp_path: Path) -> None:
    header = ["order_id", None, *VALID_HEADER[1:]]
    row = ["00123", None, *VALID_ROW[1:]]
    source = save_workbook(tmp_path / "orders.xlsx", [("Orders", [header, row])])

    with pytest.raises(XlsxStructuralError, match="header column 2 is empty"):
        load_xlsx_file(source)


def test_ignores_trailing_empty_header_cells_without_data(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [[*VALID_HEADER, None, None], VALID_ROW])],
    )

    records = load_xlsx_file(source)

    assert len(records) == 1
    assert None not in records[0]


def test_ignores_styled_trailing_empty_cells_without_real_data(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [VALID_HEADER, VALID_ROW])],
    )
    workbook = open_workbook(source)
    worksheet = workbook["Orders"]
    worksheet["J1"].number_format = "0.00"
    worksheet["J2"].number_format = "0.00"
    workbook.save(source)
    workbook.close()

    records = load_xlsx_file(source)

    assert len(records) == 1
    assert records[0]["order_id"] == "00123"


@pytest.mark.parametrize("reserved", ["source_file", "source_sheet", "source_row"])
def test_rejects_reserved_xlsx_header(tmp_path: Path, reserved: str) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [[*VALID_HEADER, reserved], [*VALID_ROW, "client value"]])],
    )

    with pytest.raises(XlsxStructuralError, match=f"reserved header columns: {reserved}"):
        load_xlsx_file(source)


def test_preserves_physically_present_blank_rows_between_records(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [
            (
                "Orders",
                [
                    VALID_HEADER,
                    VALID_ROW,
                    [None] * len(VALID_HEADER),
                    ["002", "Ada", None, "2026-08-21", "20.00", "pending"],
                ],
            )
        ],
    )

    records = load_xlsx_file(source)
    result = process_records(records)

    assert [record["source_row"] for record in records] == [2, 3, 4]
    assert records[1]["order_id"] is None
    assert result.records[1].is_invalid


def test_corrupted_xlsx_is_structural_error_with_preserved_cause(tmp_path: Path) -> None:
    source = tmp_path / "corrupted.xlsx"
    source.write_bytes(b"not an xlsx workbook")

    with pytest.raises(XlsxStructuralError) as captured:
        load_xlsx_file(source)

    assert captured.value.file_path == source
    assert captured.value.__cause__ is not None


def test_opens_workbook_without_streaming_or_cached_formula_values(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [VALID_HEADER, VALID_ROW])],
    )
    options: dict[str, object] = {}
    closed = False

    def capture_options(filename: Path, **kwargs: object):
        nonlocal closed
        options.update(kwargs)
        workbook = open_workbook(filename, **kwargs)
        original_close = workbook.close

        def track_close() -> None:
            nonlocal closed
            closed = True
            original_close()

        workbook.close = track_close
        return workbook

    monkeypatch.setattr("src.processor.load_workbook", capture_options)

    load_xlsx_file(source)

    assert options == {
        "read_only": False,
        "data_only": False,
        "keep_links": False,
    }
    assert closed


def test_closes_workbook_when_worksheet_processing_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [VALID_HEADER, VALID_ROW])],
    )
    closed = False

    def tracking_loader(filename: Path, **kwargs: object):
        nonlocal closed
        workbook = open_workbook(filename, **kwargs)
        original_close = workbook.close

        def track_close() -> None:
            nonlocal closed
            closed = True
            original_close()

        workbook.close = track_close
        return workbook

    def fail_processing(_worksheet: object, _file_path: Path):
        raise XlsxStructuralError(source, "test failure")

    monkeypatch.setattr("src.processor.load_workbook", tracking_loader)
    monkeypatch.setattr("src.processor._xlsx_header", fail_processing)

    with pytest.raises(XlsxStructuralError, match="test failure"):
        load_xlsx_file(source)

    assert closed


def test_wraps_permission_error_from_workbook_opening(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [VALID_HEADER, VALID_ROW])],
    )
    permission_error = PermissionError("access denied")

    def deny_access(*_args: object, **_kwargs: object):
        raise permission_error

    monkeypatch.setattr("src.processor.load_workbook", deny_access)

    with pytest.raises(XlsxStructuralError) as captured:
        load_xlsx_file(source)

    assert captured.value.file_path == source
    assert captured.value.__cause__ is permission_error


def test_does_not_mask_unexpected_workbook_processing_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [VALID_HEADER, VALID_ROW])],
    )

    def raise_unexpected_error(_worksheet: object, _file_path: Path):
        raise RuntimeError("unexpected bug")

    monkeypatch.setattr("src.processor._xlsx_header", raise_unexpected_error)

    with pytest.raises(RuntimeError, match="unexpected bug"):
        load_xlsx_file(source)


def test_formula_in_header_is_structural_error(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [("Orders", [['="order_id"', *VALID_HEADER[1:]], VALID_ROW])],
    )

    with pytest.raises(XlsxStructuralError, match="formula in header row"):
        load_xlsx_file(source)


@pytest.mark.parametrize("column", range(len(VALID_HEADER) + 1))
def test_formula_in_required_or_extra_data_cell_is_record_level_error(
    tmp_path: Path,
    column: int,
) -> None:
    header = [*VALID_HEADER, "extra"]
    row = [*VALID_ROW, "plain"]
    row[column] = "=1+1"
    source = save_workbook(tmp_path / "orders.xlsx", [("Orders", [header, row])])

    result = process_records(load_xlsx_file(source))

    assert len(result.invalid_records) == 1
    assert any(
        error.code == "formula_not_allowed"
        for error in result.invalid_records[0].validation_errors
    )


def test_preserves_special_xlsx_formulas_deterministically(tmp_path: Path) -> None:
    source = save_special_formula_workbook(tmp_path / "special-formulas.xlsx")
    first_records = load_xlsx_file(source)
    second_records = load_xlsx_file(source)

    first_formulas = tuple(record["note"] for record in first_records)
    second_formulas = tuple(record["note"] for record in second_records)
    assert first_formulas == second_formulas
    assert first_formulas == (
        FormulaValue("=SUM(1,2)"),
        FormulaValue("=SUM(2,3)"),
        FormulaValue(
            'dataTable(ref="G4:G4", ca="1", dt2D="1", dtr=false, '
            'r1="A1", r2="B1", del1=false, del2="1")'
        ),
    )
    assert all("object at 0x" not in formula.expression for formula in first_formulas)

    result = process_records(first_records)
    assert len(result.invalid_records) == 3
    assert all(
        any(error.code == "formula_not_allowed" for error in record.validation_errors)
        for record in result.invalid_records
    )


def test_special_xlsx_formulas_export_as_non_executable_text(tmp_path: Path) -> None:
    source = save_special_formula_workbook(tmp_path / "special-formulas.xlsx")
    result = process_records(load_xlsx_file(source))
    output = tmp_path / "sales-report.xlsx"

    export_report(result, calculate_summary(result), output)

    workbook = open_workbook(output, data_only=False)
    try:
        invalid = workbook["Invalid Records"]
        headers = [cell.value for cell in invalid[1]]
        note_column = headers.index("note") + 1
        notes = tuple(
            invalid.cell(row=row, column=note_column)
            for row in range(2, invalid.max_row + 1)
        )
        assert [cell.value for cell in notes] == [
            "'=SUM(1,2)",
            "'=SUM(2,3)",
            (
                'dataTable(ref="G4:G4", ca="1", dt2D="1", dtr=false, '
                'r1="A1", r2="B1", del1=false, del2="1")'
            ),
        ]
        assert all(cell.data_type == "s" for cell in notes)
    finally:
        workbook.close()


def test_rejects_unexpected_formula_object_without_object_repr(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(
        tmp_path / "unexpected-formula.xlsx",
        [("Orders", [[*VALID_HEADER, "note"], [*VALID_ROW, "plain"]])],
    )

    def load_with_unexpected_formula(filename: Path, **kwargs: object):
        workbook = open_workbook(filename, **kwargs)
        cell = workbook["Orders"]["G2"]
        cell._value = object()
        cell.data_type = "f"
        return workbook

    monkeypatch.setattr("src.processor.load_workbook", load_with_unexpected_formula)

    with pytest.raises(
        XlsxStructuralError,
        match="unsupported formula representation at cell G2: object",
    ) as captured:
        load_xlsx_file(source)

    assert "object at 0x" not in str(captured.value)


def test_native_xlsx_types_follow_shared_record_validation(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [
            (
                "Orders",
                [
                    VALID_HEADER,
                    ["001", "Julia", None, date(2026, 8, 20), 10, "paid"],
                    [
                        "002",
                        "Ada",
                        None,
                        datetime(2026, 8, 21, 12),
                        10.005,
                        "pending",
                    ],
                    [123, "Grace", None, date(2026, 8, 22), 0, "paid"],
                    ["004", True, True, True, True, True],
                ],
            )
        ],
    )

    result = process_records(load_xlsx_file(source))

    assert result.records[0].record["amount"] == Decimal("10.00")
    assert result.records[1].record["amount"] == Decimal("10.01")
    assert result.records[1].record["order_date"] == date(2026, 8, 21)
    assert result.records[2].is_invalid
    assert result.records[2].record["order_id"] == 123
    assert result.records[3].is_invalid


def test_combines_csv_and_xlsx_for_global_duplicate_detection(tmp_path: Path) -> None:
    csv_file = tmp_path / "a.csv"
    csv_file.write_text(
        ",".join(VALID_HEADER) + "\n00123,Julia,,2026-08-20,10.00,paid\n",
        encoding="utf-8",
    )
    save_workbook(
        tmp_path / "b.xlsx",
        [
            (
                "Orders",
                [
                    VALID_HEADER,
                    ["00123", "Ada", None, date(2026, 8, 21), 20, "pending"],
                ],
            )
        ],
    )

    records = load_supported_files(discover_supported_files(tmp_path))
    result = process_records(records)

    assert [record.record["source_file"] for record in result.duplicate_records] == [
        "a.csv",
        "b.xlsx",
    ]
    assert [record.record["source_sheet"] for record in result.records] == [
        None,
        "Orders",
    ]


def test_numeric_xlsx_id_does_not_duplicate_textual_csv_id(tmp_path: Path) -> None:
    csv_file = tmp_path / "a.csv"
    csv_file.write_text(
        ",".join(VALID_HEADER) + "\n123,Julia,,2026-08-20,10.00,paid\n",
        encoding="utf-8",
    )
    save_workbook(
        tmp_path / "b.xlsx",
        [
            (
                "Orders",
                [VALID_HEADER, [123, "Ada", None, date(2026, 8, 21), 20, "paid"]],
            )
        ],
    )

    result = process_records(
        load_supported_files(discover_supported_files(tmp_path))
    )

    assert result.duplicate_records == ()
    assert result.records[0].is_valid
    assert result.records[1].is_invalid
    assert result.records[1].record["order_id"] == 123


def test_large_scientific_amount_from_real_xlsx_is_valid(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "orders.xlsx",
        [
            (
                "Orders",
                [VALID_HEADER, ["001", "Julia", None, date(2026, 8, 20), 1e100, "paid"]],
            )
        ],
    )

    loaded = load_xlsx_file(source)
    result = process_records(loaded)

    assert loaded[0]["amount"] == 1e100
    assert result.records[0].record["amount"] == Decimal(
        "1" + ("0" * 100) + ".00"
    )
    assert result.records[0].is_valid


def test_structural_failure_in_one_format_prevents_combined_result(
    tmp_path: Path,
) -> None:
    csv_file = tmp_path / "a.csv"
    csv_file.write_text(
        ",".join(VALID_HEADER) + "\n00123,Julia,,2026-08-20,10.00,paid\n",
        encoding="utf-8",
    )
    invalid_xlsx = save_workbook(
        tmp_path / "b.xlsx",
        [("Partial", [["order_id", "status"], ["002", "paid"]])],
    )

    with pytest.raises(XlsxStructuralError, match="b.xlsx"):
        load_supported_files([csv_file, invalid_xlsx])
