from __future__ import annotations

import logging
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook

import src.exporter as exporter_module
import src.main as main_module
import src.xlsx_safety as xlsx_safety_module
from src.exporter import ReportExportError
from src.main import run


VALID_CSV = (
    "order_id,customer_name,email,order_date,amount,status\n"
    "00123,Julia,julia@example.com,2026-08-20,149.90,paid\n"
)


def test_run_returns_zero_and_logs_success(tmp_path: Path, caplog) -> None:
    (tmp_path / "orders.csv").write_text(VALID_CSV, encoding="utf-8")
    output_path = tmp_path / "output" / "sales_report.xlsx"

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, output_path)

    assert status == 0
    assert "Discovered 1 CSV input file(s)." in caplog.text
    assert "Successfully loaded 1 record(s)" in caplog.text
    assert "Processed 1 record(s): 1 valid, 0 invalid, 0 duplicate." in caplog.text
    assert "Summary: 1 total, 1 valid, 0 invalid, 0 duplicate, paid amount 149.90." in caplog.text
    assert "Successfully published Excel report" in caplog.text
    assert output_path.is_file()


def test_run_returns_nonzero_and_logs_structural_failure(
    tmp_path: Path,
    caplog,
) -> None:
    (tmp_path / "orders.csv").write_text(
        "order_id,status\n00123,paid\n", encoding="utf-8"
    )

    output_path = tmp_path / "output" / "sales_report.xlsx"
    output_path.parent.mkdir()
    output_path.write_bytes(b"previous report")

    with caplog.at_level(logging.ERROR):
        status = run(tmp_path, output_path)

    assert status == 1
    assert "Input processing failed" in caplog.text
    assert "missing required columns" in caplog.text
    assert output_path.read_bytes() == b"previous report"


def test_run_preserves_previous_report_after_prohibited_xlsx_xml(
    tmp_path: Path,
    caplog,
) -> None:
    prohibited_xml = """<!DOCTYPE worksheet [<!ENTITY safe "not-expanded">]>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData><row r="1"><c r="A1"><v>&safe;</v></c></row></sheetData>
</worksheet>
"""
    with ZipFile(tmp_path / "orders.xlsx", "w") as archive:
        archive.writestr("xl/worksheets/sheet1.xml", prohibited_xml)
    output_path = tmp_path / "output" / "sales_report.xlsx"
    output_path.parent.mkdir()
    output_path.write_bytes(b"previous report")

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, output_path)

    assert status == 1
    assert "package preflight failed" in caplog.text
    assert "Successfully published Excel report" not in caplog.text
    assert output_path.read_bytes() == b"previous report"


def test_run_controls_zip_open_permission_error_and_preserves_report(
    tmp_path: Path,
    caplog,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        ["order_id", "customer_name", "email", "order_date", "amount", "status"]
    )
    worksheet.append(["001", "Julia", None, "2026-08-20", "1.00", "paid"])
    workbook.save(tmp_path / "orders.xlsx")
    workbook.close()
    output_path = tmp_path / "output" / "sales_report.xlsx"
    output_path.parent.mkdir()
    output_path.write_bytes(b"previous report")
    permission_error = PermissionError("access denied after stat")

    def deny_zip_open(*_args: object, **_kwargs: object) -> None:
        raise permission_error

    monkeypatch.setattr(xlsx_safety_module, "ZipFile", deny_zip_open)

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, output_path)

    assert status == 1
    assert "package preflight failed" in caplog.text
    assert "Successfully loaded" not in caplog.text
    assert "Successfully published Excel report" not in caplog.text
    assert output_path.read_bytes() == b"previous report"


def test_run_returns_nonzero_when_input_directory_is_missing(
    tmp_path: Path,
    caplog,
) -> None:
    with caplog.at_level(logging.ERROR):
        status = run(
            tmp_path / "missing",
            tmp_path / "output" / "sales_report.xlsx",
        )

    assert status == 1
    assert "Input directory does not exist" in caplog.text


def test_run_returns_nonzero_when_no_supported_input_exists(
    tmp_path: Path,
    caplog,
) -> None:
    (tmp_path / "ignored.txt").write_text("ignored", encoding="utf-8")

    with caplog.at_level(logging.ERROR):
        status = run(tmp_path, tmp_path / "output" / "sales_report.xlsx")

    assert status == 1
    assert "No supported input files found" in caplog.text


def test_run_logs_and_returns_nonzero_for_filesystem_errors(
    tmp_path: Path,
    caplog,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    permission_error = PermissionError("access denied")

    def raise_permission_error(_path: Path):
        raise permission_error

    monkeypatch.setattr(Path, "iterdir", raise_permission_error)

    with caplog.at_level(logging.ERROR):
        status = run(tmp_path, tmp_path / "output" / "sales_report.xlsx")

    assert status == 1
    assert "Input processing failed" in caplog.text
    assert "Could not access input directory" in caplog.text


def test_run_does_not_mask_unexpected_programming_errors(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def raise_programming_error(_path: Path):
        raise RuntimeError("unexpected bug")

    monkeypatch.setattr(Path, "iterdir", raise_programming_error)

    with pytest.raises(RuntimeError, match="unexpected bug"):
        run(tmp_path, tmp_path / "output" / "sales_report.xlsx")


def test_run_succeeds_and_logs_record_level_validation_failures(
    tmp_path: Path,
    caplog,
) -> None:
    (tmp_path / "orders.csv").write_text(
        "order_id,customer_name,email,order_date,amount,status\n"
        ",,invalid,01/02/2026,-1.00,unknown\n",
        encoding="utf-8",
    )

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, tmp_path / "output" / "sales_report.xlsx")

    assert status == 0
    assert "Processed 1 record(s): 0 valid, 1 invalid, 0 duplicate." in caplog.text


def test_run_classifies_duplicates_across_multiple_csv_files(
    tmp_path: Path,
    caplog,
) -> None:
    (tmp_path / "a.csv").write_text(VALID_CSV, encoding="utf-8")
    (tmp_path / "b.csv").write_text(VALID_CSV, encoding="utf-8")

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, tmp_path / "output" / "sales_report.xlsx")

    assert status == 0
    assert "Processed 2 record(s): 0 valid, 0 invalid, 2 duplicate." in caplog.text


def test_run_processes_csv_and_xlsx_together(tmp_path: Path, caplog) -> None:
    (tmp_path / "a.csv").write_text(VALID_CSV, encoding="utf-8")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders"
    worksheet.append(
        ["order_id", "customer_name", "email", "order_date", "amount", "status"]
    )
    worksheet.append(["00123", "Ada", None, "2026-08-21", 20, "pending"])
    workbook.save(tmp_path / "b.xlsx")
    workbook.close()

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, tmp_path / "output" / "sales_report.xlsx")

    assert status == 0
    assert "Discovered 1 CSV input file(s)." in caplog.text
    assert "Discovered 1 XLSX input file(s)." in caplog.text
    assert "Processed 2 record(s): 0 valid, 0 invalid, 2 duplicate." in caplog.text


def test_run_returns_nonzero_for_structurally_invalid_xlsx(
    tmp_path: Path,
    caplog,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["order_id", "status"])
    worksheet.append(["00123", "paid"])
    workbook.save(tmp_path / "orders.xlsx")
    workbook.close()

    with caplog.at_level(logging.ERROR):
        status = run(tmp_path, tmp_path / "output" / "sales_report.xlsx")

    assert status == 1
    assert "XLSX structural error" in caplog.text
    assert "missing required columns" in caplog.text


def test_run_does_not_calculate_summary_after_structural_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (tmp_path / "invalid.csv").write_text(
        "order_id,status\n00123,paid\n",
        encoding="utf-8",
    )

    def unexpected_summary_calculation(_result: object) -> None:
        pytest.fail("summary must not be calculated after structural failure")

    monkeypatch.setattr(
        main_module,
        "calculate_summary",
        unexpected_summary_calculation,
    )

    assert run(tmp_path, tmp_path / "output" / "sales_report.xlsx") == 1


def test_run_returns_nonzero_and_does_not_log_success_on_export_failure(
    tmp_path: Path,
    caplog,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (tmp_path / "orders.csv").write_text(VALID_CSV, encoding="utf-8")
    output_path = tmp_path / "output" / "sales_report.xlsx"
    output_path.parent.mkdir()
    output_path.write_bytes(b"previous report")

    def fail_export(*_args: object) -> None:
        cause = PermissionError("access denied")
        raise ReportExportError(output_path, "publication failed") from cause

    monkeypatch.setattr(main_module, "export_report", fail_export)

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, output_path)

    assert status == 1
    assert "Report export failed" in caplog.text
    assert "Successfully published Excel report" not in caplog.text
    assert output_path.read_bytes() == b"previous report"


def test_run_does_not_log_success_for_logically_incomplete_saved_workbook(
    tmp_path: Path,
    caplog,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (tmp_path / "orders.csv").write_text(VALID_CSV, encoding="utf-8")
    output_path = tmp_path / "output" / "sales_report.xlsx"
    output_path.parent.mkdir()
    output_path.write_bytes(b"previous report")
    real_save = Workbook.save

    def save_empty_workbook(workbook: Workbook, path: Path) -> None:
        real_save(workbook, path)
        corrupted = load_workbook(path, data_only=False)
        try:
            for worksheet in corrupted.worksheets:
                worksheet.delete_rows(1, worksheet.max_row)
            real_save(corrupted, path)
        finally:
            corrupted.close()

    monkeypatch.setattr(Workbook, "save", save_empty_workbook)

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, output_path)

    assert status == 1
    assert "saved worksheet 'Valid Records' has an invalid projection count" in caplog.text
    assert "Successfully published Excel report" not in caplog.text
    assert output_path.read_bytes() == b"previous report"


def test_run_does_not_log_success_when_temporary_creation_fails(
    tmp_path: Path,
    caplog,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (tmp_path / "orders.csv").write_text(VALID_CSV, encoding="utf-8")
    output_path = tmp_path / "output" / "sales_report.xlsx"
    output_path.parent.mkdir()
    output_path.write_bytes(b"previous report")
    cause = OSError("temporary creation failed")

    def fail_temporary_creation(*_args: object, **_kwargs: object) -> tuple[int, str]:
        raise cause

    monkeypatch.setattr(exporter_module.tempfile, "mkstemp", fail_temporary_creation)

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, output_path)

    assert status == 1
    assert "output location could not be prepared" in caplog.text
    assert "Successfully published Excel report" not in caplog.text
    assert output_path.read_bytes() == b"previous report"


def test_run_does_not_log_success_for_shared_money_serializer_failure(
    tmp_path: Path,
    caplog,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (tmp_path / "orders.csv").write_text(VALID_CSV, encoding="utf-8")
    output_path = tmp_path / "output" / "sales_report.xlsx"
    output_path.parent.mkdir()
    output_path.write_bytes(b"previous report")

    monkeypatch.setattr(
        exporter_module,
        "_money_text",
        lambda _value, _output_path: "0.00",
    )

    with caplog.at_level(logging.INFO):
        status = run(tmp_path, output_path)

    assert status == 1
    assert "logical summary differs from the supplied summary" in caplog.text
    assert "Successfully published Excel report" not in caplog.text
    assert output_path.read_bytes() == b"previous report"


def test_run_does_not_mask_unexpected_export_errors(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    (tmp_path / "orders.csv").write_text(VALID_CSV, encoding="utf-8")

    def fail_unexpectedly(*_args: object) -> None:
        raise RuntimeError("unexpected export bug")

    monkeypatch.setattr(main_module, "export_report", fail_unexpectedly)

    with pytest.raises(RuntimeError, match="unexpected export bug"):
        run(tmp_path, tmp_path / "output" / "sales_report.xlsx")
