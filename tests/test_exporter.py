from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import src.exporter as exporter_module
from src.exporter import (
    DUPLICATES_SHEET,
    INVALID_SHEET,
    KNOWN_COLUMNS,
    SUMMARY_SHEET,
    TRACEABILITY_COLUMNS,
    VALID_SHEET,
    WORKSHEET_NAMES,
    ReportExportError,
    _validate_text,
    _validate_worksheet_size,
    export_report,
)
from src.processor import (
    ProcessedRecord,
    ProcessingResult,
    discover_supported_files,
    load_supported_files,
    process_records,
)
from src.summary import ProcessingSummary, calculate_summary
from src.validator import FormulaValue, ValidationError


def raw_record(order_id: str, **overrides: object) -> dict[str, object]:
    record: dict[str, object] = {
        "order_id": order_id,
        "customer_name": "Julia",
        "email": None,
        "order_date": "2026-08-20",
        "amount": "10.00",
        "status": "paid",
        "source_file": "orders.csv",
        "source_sheet": None,
        "source_row": 2,
    }
    record.update(overrides)
    return record


def logical_workbook(path: Path) -> tuple[tuple[str, tuple[tuple[object, ...], ...]], ...]:
    workbook = load_workbook(path, data_only=False)
    try:
        return tuple(
            (
                worksheet.title,
                tuple(
                    tuple(cell.value for cell in row)
                    for row in worksheet.iter_rows()
                ),
            )
            for worksheet in workbook.worksheets
        )
    finally:
        workbook.close()


def logical_formatting(
    path: Path,
) -> tuple[tuple[str, str | None, str | None, tuple[tuple[str, ...], ...]], ...]:
    workbook = load_workbook(path, data_only=False)
    try:
        return tuple(
            (
                worksheet.title,
                worksheet.freeze_panes,
                worksheet.auto_filter.ref,
                tuple(
                    tuple(cell.number_format for cell in row)
                    for row in worksheet.iter_rows()
                ),
            )
            for worksheet in workbook.worksheets
        )
    finally:
        workbook.close()


def temp_reports(output_path: Path) -> list[Path]:
    return list(output_path.parent.glob(f".{output_path.name}.*.xlsx"))


def test_creates_exact_workbook_structure_and_uses_supplied_summary(
    tmp_path: Path,
) -> None:
    result = process_records([raw_record("001")])
    supplied_summary = ProcessingSummary(
        total_records=91,
        valid_records=81,
        invalid_records=7,
        duplicate_records=5,
        total_paid_amount=Decimal("123456789012345678901234567890.12"),
    )
    output_path = tmp_path / "output" / "sales_report.xlsx"

    export_report(result, supplied_summary, output_path)

    assert temp_reports(output_path) == []
    workbook = load_workbook(output_path, data_only=False)
    try:
        assert tuple(workbook.sheetnames) == WORKSHEET_NAMES
        assert all(not worksheet.merged_cells for worksheet in workbook.worksheets)
        assert all(not worksheet._charts for worksheet in workbook.worksheets)
        summary = workbook[SUMMARY_SHEET]
        assert summary["A1"].font.bold
        assert summary["B1"].font.bold
        assert [summary.cell(row=row, column=1).value for row in range(2, 7)] == [
            "Total Records",
            "Valid Records",
            "Invalid Records",
            "Duplicate Records",
            "Total Paid Amount (USD)",
        ]
        assert [summary.cell(row=row, column=2).value for row in range(2, 6)] == [
            91,
            81,
            7,
            5,
        ]
        assert summary["B6"].value == "123456789012345678901234567890.12"
        assert summary["B6"].data_type == "s"
        assert summary["A8"].value == (
            "Invalid and duplicate record counts may overlap."
        )
    finally:
        workbook.close()


def test_preserves_projections_columns_values_and_validation_errors(
    tmp_path: Path,
) -> None:
    result = process_records(
        [
            raw_record(
                "00123",
                customer_name="Ada",
                order_date="2026-08-20",
                amount="149.90",
                source_file="a.csv",
                source_row=2,
                zeta="online",
                validation_errors="client value",
                validation_errors_2="second client value",
            ),
            raw_record(
                "invalid",
                email="bad",
                source_file="b.xlsx",
                source_sheet="Orders",
                source_row=4,
                Alpha="store",
            ),
            raw_record(
                "duplicate",
                source_file="a.csv",
                source_row=3,
            ),
            raw_record(
                "duplicate",
                status="unknown",
                source_file="b.xlsx",
                source_sheet="Orders",
                source_row=5,
            ),
        ]
    )
    output_path = tmp_path / "sales_report.xlsx"

    export_report(result, calculate_summary(result), output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        expected_base = [
            *KNOWN_COLUMNS,
            "Alpha",
            "validation_errors",
            "validation_errors_2",
            "zeta",
            *TRACEABILITY_COLUMNS,
        ]
        valid = workbook[VALID_SHEET]
        invalid = workbook[INVALID_SHEET]
        duplicates = workbook[DUPLICATES_SHEET]

        assert [cell.value for cell in valid[1]] == expected_base
        assert all(cell.font.bold for cell in valid[1])
        assert [cell.value for cell in invalid[1]] == [
            *expected_base,
            "validation_errors_3",
        ]
        assert [cell.value for cell in duplicates[1]] == [
            *expected_base,
            "validation_errors_3",
        ]
        assert valid.max_row == 2
        assert invalid.max_row == 3
        assert duplicates.max_row == 3

        valid_values = [cell.value for cell in valid[2]]
        assert valid_values[0] == "00123"
        assert valid["A2"].data_type == "s"
        assert valid_values[3] == date(2026, 8, 20)
        assert valid["D2"].number_format == "yyyy-mm-dd"
        assert valid_values[4] == "149.90"
        assert valid["E2"].data_type == "s"
        assert valid_values[6:10] == [None, "client value", "second client value", "online"]
        assert valid_values[-3:] == ["a.csv", None, 2]

        invalid_error_column = invalid.max_column
        invalid_errors = [
            invalid.cell(row=row, column=invalid_error_column).value
            for row in range(2, invalid.max_row + 1)
        ]
        assert invalid_errors == [
            "email [invalid_email]: email format is invalid",
            "status [invalid_status]: status must be one of: paid, pending, cancelled, refunded",
        ]
        duplicate_errors = [
            duplicates.cell(row=row, column=duplicates.max_column).value
            for row in range(2, duplicates.max_row + 1)
        ]
        assert duplicate_errors == [
            None,
            "status [invalid_status]: status must be one of: paid, pending, cancelled, refunded",
        ]
        assert [duplicates.cell(row=row, column=1).value for row in (2, 3)] == [
            "duplicate",
            "duplicate",
        ]
    finally:
        workbook.close()


def test_monetary_values_survive_round_trip_as_exact_two_place_text(
    tmp_path: Path,
) -> None:
    large_amount = "123456789012345678901234567890.12"
    result = process_records(
        [
            raw_record("zero", amount="0", status="pending", source_row=2),
            raw_record("normal", amount="149.90", status="pending", source_row=3),
            raw_record("large", amount=large_amount, status="paid", source_row=4),
        ]
    )
    output_path = tmp_path / "sales_report.xlsx"

    export_report(result, calculate_summary(result), output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        assert [workbook[VALID_SHEET].cell(row=row, column=5).value for row in (2, 3, 4)] == [
            "0.00",
            "149.90",
            large_amount,
        ]
        assert workbook[SUMMARY_SHEET]["B6"].value == large_amount
        assert all(
            workbook[VALID_SHEET].cell(row=row, column=5).data_type == "s"
            for row in (2, 3, 4)
        )
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("value", "expected_record", "expected_summary"),
    [
        (Decimal("0.00"), "0.00", "0.00"),
        (Decimal("-0.00"), "'-0.00", "-0.00"),
        (Decimal("0.01"), "0.01", "0.01"),
        (Decimal("123.45"), "123.45", "123.45"),
        (
            Decimal("999999999999999999999999999999.99"),
            "999999999999999999999999999999.99",
            "999999999999999999999999999999.99",
        ),
    ],
)
def test_independent_money_invariant_preserves_normalized_decimal_round_trip(
    tmp_path: Path,
    value: Decimal,
    expected_record: str,
    expected_summary: str,
) -> None:
    record = ProcessedRecord(
        record=raw_record("001", amount=value),
        validation_errors=(),
        is_duplicate=False,
    )
    result = ProcessingResult((record,), (record,), (), ())
    output_path = tmp_path / "sales_report.xlsx"

    export_report(
        result,
        ProcessingSummary(1, 1, 0, 0, value),
        output_path,
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        assert workbook[VALID_SHEET]["E2"].value == expected_record
        assert workbook[SUMMARY_SHEET]["B6"].value == expected_summary
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "value",
    [
        Decimal("1.234"),
        Decimal("NaN"),
        Decimal("Infinity"),
        Decimal("-Infinity"),
    ],
)
def test_independent_money_invariant_rejects_unrepresentable_decimal(
    tmp_path: Path,
    value: Decimal,
) -> None:
    record = ProcessedRecord(
        record=raw_record("001", amount=value),
        validation_errors=(),
        is_duplicate=False,
    )
    result = ProcessingResult((record,), (record,), (), ())
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")

    with pytest.raises(ReportExportError):
        export_report(
            result,
            ProcessingSummary(1, 1, 0, 0, Decimal("0.00")),
            output_path,
        )

    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


@pytest.mark.parametrize(
    "corruption",
    ["always_zero", "change_cents", "remove_trailing_zero", "change_sign"],
)
def test_independent_money_invariant_rejects_shared_serializer_corruption(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    corruption: str,
) -> None:
    record_amount = (
        Decimal("123.40")
        if corruption == "remove_trailing_zero"
        else Decimal("123.45")
    )
    record = ProcessedRecord(
        record=raw_record("001", amount=record_amount),
        validation_errors=(),
        is_duplicate=False,
    )
    result = ProcessingResult((record,), (record,), (), ())
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    real_money_text = exporter_module._money_text

    def corrupt_money_text(value: Decimal, output_path: Path) -> str:
        correct = real_money_text(value, output_path)
        if corruption == "always_zero":
            return "0.00"
        if corruption == "change_cents":
            return f"{correct[:-2]}99"
        if corruption == "remove_trailing_zero":
            return correct.rstrip("0")
        if corruption == "change_sign":
            return f"-{correct}"
        pytest.fail(f"unknown monetary corruption: {corruption}")

    monkeypatch.setattr(exporter_module, "_money_text", corrupt_money_text)

    def fail_if_replace_is_attempted(_source: Path, _destination: Path) -> None:
        pytest.fail("os.replace must not run after monetary invariant failure")

    monkeypatch.setattr(exporter_module.os, "replace", fail_if_replace_is_attempted)

    with pytest.raises(ReportExportError):
        export_report(
            result,
            ProcessingSummary(1, 1, 0, 0, Decimal("987.65")),
            output_path,
        )

    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


@pytest.mark.parametrize("target", ["record", "summary"])
def test_record_and_summary_money_have_independent_common_mode_protection(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    target: str,
) -> None:
    record_amount = Decimal("123.45")
    summary_amount = Decimal("987.65")
    record = ProcessedRecord(
        record=raw_record("001", amount=record_amount),
        validation_errors=(),
        is_duplicate=False,
    )
    result = ProcessingResult((record,), (record,), (), ())
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    real_money_text = exporter_module._money_text

    def corrupt_selected_money(value: Decimal, output_path: Path) -> str:
        if (target == "record" and value == record_amount) or (
            target == "summary" and value == summary_amount
        ):
            return "0.00"
        return real_money_text(value, output_path)

    monkeypatch.setattr(
        exporter_module,
        "_money_text",
        corrupt_selected_money,
    )

    def fail_if_replace_is_attempted(_source: Path, _destination: Path) -> None:
        pytest.fail("os.replace must not run after monetary invariant failure")

    monkeypatch.setattr(exporter_module.os, "replace", fail_if_replace_is_attempted)

    with pytest.raises(ReportExportError):
        export_report(
            result,
            ProcessingSummary(1, 1, 0, 0, summary_amount),
            output_path,
        )

    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_exporter_uses_projection_membership_without_recalculation(
    tmp_path: Path,
) -> None:
    record = ProcessedRecord(
        record=raw_record("001"),
        validation_errors=(),
        is_duplicate=False,
    )
    result = ProcessingResult(
        records=(record,),
        valid_records=(),
        invalid_records=(record,),
        duplicate_records=(record,),
    )
    output_path = tmp_path / "sales_report.xlsx"

    export_report(
        result,
        ProcessingSummary(1, 0, 1, 1, Decimal("0.00")),
        output_path,
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        assert workbook[VALID_SHEET].max_row == 1
        assert workbook[INVALID_SHEET]["A2"].value == "001"
        assert workbook[DUPLICATES_SHEET]["A2"].value == "001"
    finally:
        workbook.close()


def test_rejects_processed_record_missing_from_every_projection(
    tmp_path: Path,
) -> None:
    record = ProcessedRecord(
        record=raw_record("001"),
        validation_errors=(),
        is_duplicate=False,
    )
    result = ProcessingResult(
        records=(record,),
        valid_records=(),
        invalid_records=(),
        duplicate_records=(),
    )
    output_path = tmp_path / "sales_report.xlsx"

    with pytest.raises(ReportExportError, match="absent from every output projection"):
        export_report(
            result,
            ProcessingSummary(1, 0, 0, 0, Decimal("0.00")),
            output_path,
        )

    assert not output_path.exists()


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("=SUM(A1:A2)", "'=SUM(A1:A2)"),
        ("+cmd", "'+cmd"),
        ("-attack", "'-attack"),
        ("@example", "'@example"),
        (" =SUM(A1:A2)", "' =SUM(A1:A2)"),
        ("\t+cmd", "'\t+cmd"),
        ("\r\n-attack", "'\r\n-attack"),
    ],
)
def test_neutralizes_formula_triggering_input_text(
    tmp_path: Path,
    value: str,
    expected: str,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"

    result = process_records([raw_record("001", payload=value)])
    export_report(result, calculate_summary(result), output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        worksheet = workbook[VALID_SHEET]
        headers = [cell.value for cell in worksheet[1]]
        cell = worksheet.cell(row=2, column=headers.index("payload") + 1)
        assert cell.value == expected
        assert cell.data_type != "f"
    finally:
        workbook.close()


def test_neutralizes_formula_values_extra_names_and_source_metadata(
    tmp_path: Path,
) -> None:
    result = process_records(
        [
            raw_record(
                "001",
                source_file="-orders.xlsx",
                source_sheet="\t=Orders",
                **{
                    "=dangerous_header": "+payload",
                    "formula_input": FormulaValue("=1+1"),
                    "at_value": "@payload",
                },
            )
        ]
    )
    output_path = tmp_path / "sales_report.xlsx"

    export_report(result, calculate_summary(result), output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        invalid = workbook[INVALID_SHEET]
        headers = [cell.value for cell in invalid[1]]
        dangerous_header_column = headers.index("'=dangerous_header") + 1
        formula_column = headers.index("formula_input") + 1
        at_column = headers.index("at_value") + 1
        source_file_column = headers.index("source_file") + 1
        source_sheet_column = headers.index("source_sheet") + 1

        assert invalid.cell(2, dangerous_header_column).value == "'+payload"
        assert invalid.cell(2, formula_column).value == "'=1+1"
        assert invalid.cell(2, at_column).value == "'@payload"
        assert invalid.cell(2, source_file_column).value == "'-orders.xlsx"
        assert invalid.cell(2, source_sheet_column).value == "'\t=Orders"
        assert all(
            cell.data_type != "f"
            for sheet_name in (VALID_SHEET, INVALID_SHEET, DUPLICATES_SHEET)
            for cell in workbook[sheet_name][1]
        )
        assert all(cell.data_type != "f" for cell in invalid[2])
    finally:
        workbook.close()


def test_neutralizes_formula_trigger_in_known_field(tmp_path: Path) -> None:
    result = process_records([raw_record("001", customer_name="=DANGEROUS()")])
    output_path = tmp_path / "sales_report.xlsx"

    export_report(result, calculate_summary(result), output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        cell = workbook[VALID_SHEET]["B2"]
        assert cell.value == "'=DANGEROUS()"
        assert cell.data_type != "f"
    finally:
        workbook.close()


def test_rejects_column_collision_created_by_formula_protection(
    tmp_path: Path,
) -> None:
    result = process_records(
        [raw_record("001", **{"=header": "first", "'=header": "second"})]
    )
    output_path = tmp_path / "sales_report.xlsx"

    with pytest.raises(ReportExportError, match="collide after formula-injection"):
        export_report(result, calculate_summary(result), output_path)

    assert not output_path.exists()


def test_rejects_empty_processed_record_column_name(tmp_path: Path) -> None:
    result = process_records([raw_record("001", **{"": "untraceable"})])
    output_path = tmp_path / "sales_report.xlsx"

    with pytest.raises(ReportExportError, match="column names must not be empty"):
        export_report(result, calculate_summary(result), output_path)

    assert not output_path.exists()


def test_rejects_unsupported_values_without_publishing(tmp_path: Path) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    previous_content = b"previous report"
    output_path.write_bytes(previous_content)
    result = process_records([raw_record("001", unsupported=object())])

    with pytest.raises(ReportExportError, match="unsupported cell value type") as caught:
        export_report(result, calculate_summary(result), output_path)

    assert isinstance(caught.value.__cause__, ValueError)
    assert output_path.read_bytes() == previous_content
    assert temp_reports(output_path) == []


def test_rejects_invalid_summary_types_without_recalculation(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    result = process_records([raw_record("001")])
    invalid_summary = ProcessingSummary(
        total_records="1",  # type: ignore[arg-type]
        valid_records=1,
        invalid_records=0,
        duplicate_records=0,
        total_paid_amount=Decimal("10.00"),
    )

    with pytest.raises(ReportExportError, match="must be a non-negative integer"):
        export_report(result, invalid_summary, output_path)

    assert not output_path.exists()


def test_rejects_non_textual_validation_error_parts(tmp_path: Path) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    record = ProcessedRecord(
        record=raw_record("001"),
        validation_errors=(
            ValidationError(
                field=1,  # type: ignore[arg-type]
                code="invalid",
                message="invalid field type",
            ),
        ),
        is_duplicate=False,
    )
    result = ProcessingResult(
        records=(record,),
        valid_records=(),
        invalid_records=(record,),
        duplicate_records=(),
    )

    with pytest.raises(ReportExportError, match="validation errors must contain"):
        export_report(
            result,
            ProcessingSummary(1, 0, 1, 0, Decimal("0.00")),
            output_path,
        )

    assert not output_path.exists()


def test_rejects_xml_incompatible_text_without_publishing(tmp_path: Path) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    result = process_records([raw_record("001", note="unsafe\x01value")])

    with pytest.raises(ReportExportError, match="XML-incompatible"):
        export_report(result, calculate_summary(result), output_path)

    assert not output_path.exists()
    assert temp_reports(output_path) == []


def test_rejects_float_that_openpyxl_cannot_round_trip_exactly(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    value = 1.2345678901234567
    result = process_records([raw_record("001", measured_value=value)])

    with pytest.raises(ReportExportError, match="exact XLSX numeric serialization"):
        export_report(result, calculate_summary(result), output_path)

    assert not output_path.exists()


def test_allows_positive_float_zero_round_trip(tmp_path: Path) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    result = process_records([raw_record("001", measured_value=0.0)])

    export_report(result, calculate_summary(result), output_path)

    workbook = load_workbook(output_path, data_only=False)
    try:
        headers = [cell.value for cell in workbook[VALID_SHEET][1]]
        value = workbook[VALID_SHEET].cell(
            row=2,
            column=headers.index("measured_value") + 1,
        ).value
        assert value == 0
    finally:
        workbook.close()


def test_rejects_negative_signed_float_zero_without_replacing_report(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    result = process_records([raw_record("001", measured_value=-0.0)])

    with pytest.raises(ReportExportError, match="negative signed float zero"):
        export_report(result, calculate_summary(result), output_path)

    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_rejects_sub_millisecond_datetime_precision(tmp_path: Path) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    result = process_records(
        [
            raw_record(
                "001",
                captured_at=datetime(2026, 8, 21, 12, 30, 0, 123456),
            )
        ]
    )

    with pytest.raises(ReportExportError, match="finer than one millisecond"):
        export_report(result, calculate_summary(result), output_path)

    assert not output_path.exists()


def test_validates_cell_and_worksheet_limits_without_large_files(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"

    _validate_text("a" * 32_767, output_path)
    with pytest.raises(ReportExportError, match="cell text exceeds"):
        _validate_text("a" * 32_768, output_path)
    with pytest.raises(ReportExportError, match="cell text exceeds"):
        exporter_module._protect_text("=" + ("a" * 32_766), output_path)
    _validate_worksheet_size(1_048_576, 16_384, output_path, "Records")
    with pytest.raises(ReportExportError, match="row limit"):
        _validate_worksheet_size(1_048_577, 1, output_path, "Records")
    with pytest.raises(ReportExportError, match="column limit"):
        _validate_worksheet_size(1, 16_385, output_path, "Records")


def test_replaces_existing_report_only_after_success(tmp_path: Path) -> None:
    output_path = tmp_path / "output" / "sales_report.xlsx"
    output_path.parent.mkdir()
    output_path.write_bytes(b"previous report")
    result = process_records([raw_record("001")])

    export_report(result, calculate_summary(result), output_path)

    assert output_path.read_bytes() != b"previous report"
    assert logical_workbook(output_path)[0][0] == SUMMARY_SHEET
    assert temp_reports(output_path) == []


def test_build_failure_preserves_existing_report_and_cause(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    cause = ValueError("construction failed")

    def fail_build(*_args: object) -> Workbook:
        raise ReportExportError(output_path, "construction failed") from cause

    monkeypatch.setattr(exporter_module, "_build_logical_report", fail_build)

    with pytest.raises(ReportExportError) as caught:
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert caught.value.__cause__ is cause
    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_save_failure_preserves_existing_report_and_removes_temporary_file(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    cause = PermissionError("save denied")

    def fail_save(_workbook: Workbook, _path: Path) -> None:
        raise cause

    monkeypatch.setattr(Workbook, "save", fail_save)

    with pytest.raises(ReportExportError, match="could not be saved") as caught:
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert caught.value.__cause__ is cause
    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_invalid_saved_workbook_is_not_published(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")

    def save_invalid_workbook(_workbook: Workbook, path: Path) -> None:
        Path(path).write_bytes(b"not an XLSX archive")

    monkeypatch.setattr(Workbook, "save", save_invalid_workbook)

    with pytest.raises(ReportExportError, match="temporary report"):
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


@pytest.mark.parametrize(
    "corruption",
    [
        "empty_workbook",
        "summary_value",
        "record_header",
        "omitted_record",
        "monetary_value",
        "record_order",
        "validation_errors",
    ],
)
def test_logically_incorrect_saved_workbook_is_not_published(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    corruption: str,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    result = process_records(
        [
            raw_record("001", source_row=2),
            raw_record("002", amount="20.00", source_row=3),
            raw_record("invalid", email="bad", source_row=4),
        ]
    )
    real_save = Workbook.save

    def save_corrupted_workbook(workbook: Workbook, path: Path) -> None:
        real_save(workbook, path)
        corrupted = load_workbook(path, data_only=False)
        try:
            if corruption == "empty_workbook":
                for worksheet in corrupted.worksheets:
                    worksheet.delete_rows(1, worksheet.max_row)
            elif corruption == "summary_value":
                corrupted[SUMMARY_SHEET]["B2"] = 999
            elif corruption == "record_header":
                corrupted[VALID_SHEET]["A1"] = "changed_order_id"
            elif corruption == "omitted_record":
                corrupted[VALID_SHEET].delete_rows(3)
            elif corruption == "monetary_value":
                corrupted[SUMMARY_SHEET]["B6"] = "999.99"
            elif corruption == "record_order":
                worksheet = corrupted[VALID_SHEET]
                for column_index in range(1, worksheet.max_column + 1):
                    first = worksheet.cell(2, column_index).value
                    second = worksheet.cell(3, column_index).value
                    worksheet.cell(2, column_index).value = second
                    worksheet.cell(3, column_index).value = first
            elif corruption == "validation_errors":
                worksheet = corrupted[INVALID_SHEET]
                worksheet.cell(2, worksheet.max_column).value = None
            else:
                pytest.fail(f"unknown corruption: {corruption}")
            real_save(corrupted, path)
        finally:
            corrupted.close()

    monkeypatch.setattr(Workbook, "save", save_corrupted_workbook)

    with pytest.raises(
        ReportExportError,
        match="saved|validation failed",
    ) as caught:
        export_report(result, calculate_summary(result), output_path)

    assert isinstance(caught.value.__cause__, ValueError)
    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


@pytest.mark.parametrize(
    "renderer_bug",
    ["omit_record", "omit_header", "alter_header", "reverse_rows"],
)
def test_renderer_bug_cannot_change_independent_logical_expectation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    renderer_bug: str,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    result = process_records(
        [
            raw_record("001", source_row=2),
            raw_record("002", amount="20.00", source_row=3),
        ]
    )
    real_renderer = exporter_module._render_logical_workbook

    def render_incorrectly(model: object) -> Workbook:
        workbook = real_renderer(model)
        worksheet = workbook[VALID_SHEET]
        if renderer_bug == "omit_record":
            worksheet.delete_rows(3)
        elif renderer_bug == "omit_header":
            for cell in worksheet[1]:
                cell.value = None
        elif renderer_bug == "alter_header":
            worksheet["A1"] = "changed_order_id"
        elif renderer_bug == "reverse_rows":
            for column_index in range(1, worksheet.max_column + 1):
                first = worksheet.cell(2, column_index).value
                second = worksheet.cell(3, column_index).value
                worksheet.cell(2, column_index).value = second
                worksheet.cell(3, column_index).value = first
        else:
            pytest.fail(f"unknown renderer bug: {renderer_bug}")
        return workbook

    monkeypatch.setattr(
        exporter_module,
        "_render_logical_workbook",
        render_incorrectly,
    )

    with pytest.raises(ReportExportError):
        export_report(result, calculate_summary(result), output_path)

    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_formula_safety_invariant_is_independent_from_text_protection(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    result = process_records([raw_record("001", payload="=1+1")])

    monkeypatch.setattr(
        exporter_module,
        "_protect_text",
        lambda value, _output_path: value,
    )

    with pytest.raises(ReportExportError, match="was not neutralized"):
        export_report(result, calculate_summary(result), output_path)

    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_logical_projection_counts_come_directly_from_processing_result(
    tmp_path: Path,
) -> None:
    result = process_records(
        [
            raw_record("valid"),
            raw_record("invalid", email="bad"),
            raw_record("duplicate", source_row=3),
            raw_record("duplicate", source_row=4),
        ]
    )

    model = exporter_module._build_logical_report(
        result,
        calculate_summary(result),
        tmp_path / "sales_report.xlsx",
    )

    assert len(model[1].cells) - 1 == len(result.valid_records)
    assert len(model[2].cells) - 1 == len(result.invalid_records)
    assert len(model[3].cells) - 1 == len(result.duplicate_records)


def test_temporary_creation_failure_preserves_report_and_cause(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    cause = OSError("temporary creation failed")

    def fail_temporary_creation(*_args: object, **_kwargs: object) -> tuple[int, str]:
        raise cause

    monkeypatch.setattr(exporter_module.tempfile, "mkstemp", fail_temporary_creation)

    with pytest.raises(ReportExportError, match="could not be prepared") as caught:
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert caught.value.__cause__ is cause
    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_close_failure_prevents_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    cause = PermissionError("close denied")

    def fail_close(_workbook: Workbook) -> None:
        raise cause

    monkeypatch.setattr(Workbook, "close", fail_close)

    with pytest.raises(ReportExportError, match="could not be closed") as caught:
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert caught.value.__cause__ is cause
    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_replace_failure_preserves_existing_report_and_cause(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.write_bytes(b"previous report")
    cause = PermissionError("replace denied")
    real_replace = exporter_module.os.replace

    def fail_replace(source: Path, destination: Path) -> None:
        if Path(destination) == output_path:
            raise cause
        real_replace(source, destination)

    monkeypatch.setattr(exporter_module.os, "replace", fail_replace)

    with pytest.raises(ReportExportError, match="atomically") as caught:
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert caught.value.__cause__ is cause
    assert output_path.read_bytes() == b"previous report"
    assert temp_reports(output_path) == []


def test_rejects_final_path_that_is_a_directory(tmp_path: Path) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    output_path.mkdir()

    with pytest.raises(ReportExportError, match="output path is a directory") as caught:
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert isinstance(caught.value.__cause__, IsADirectoryError)


def test_wraps_output_directory_permission_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "output" / "sales_report.xlsx"
    cause = PermissionError("directory denied")

    def fail_mkdir(*_args: object, **_kwargs: object) -> None:
        raise cause

    monkeypatch.setattr(Path, "mkdir", fail_mkdir)

    with pytest.raises(ReportExportError, match="could not be prepared") as caught:
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert caught.value.__cause__ is cause


def test_cleanup_failure_is_attached_without_masking_primary_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"
    primary_cause = PermissionError("save denied")
    cleanup_cause = PermissionError("cleanup denied")

    def fail_save(_workbook: Workbook, _path: Path) -> None:
        raise primary_cause

    def fail_unlink(*_args: object, **_kwargs: object) -> None:
        raise cleanup_cause

    monkeypatch.setattr(Workbook, "save", fail_save)
    monkeypatch.setattr(Path, "unlink", fail_unlink)

    with pytest.raises(ReportExportError) as caught:
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )

    assert caught.value.__cause__ is primary_cause
    assert any("cleanup also failed" in note for note in caught.value.__notes__)


def test_unexpected_build_error_is_not_masked(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = tmp_path / "sales_report.xlsx"

    def fail_unexpectedly(*_args: object) -> Workbook:
        raise RuntimeError("unexpected bug")

    monkeypatch.setattr(exporter_module, "_build_logical_report", fail_unexpectedly)

    with pytest.raises(RuntimeError, match="unexpected bug"):
        export_report(
            process_records([raw_record("001")]),
            ProcessingSummary(1, 1, 0, 0, Decimal("10.00")),
            output_path,
        )


def test_logical_output_is_deterministic(tmp_path: Path) -> None:
    result = process_records(
        [
            raw_record("001", zeta="last", Alpha="first"),
            raw_record("002", email="invalid", Beta="second"),
        ]
    )
    summary = calculate_summary(result)
    first_path = tmp_path / "first.xlsx"
    second_path = tmp_path / "second.xlsx"

    export_report(result, summary, first_path)
    export_report(result, summary, second_path)

    assert logical_workbook(first_path) == logical_workbook(second_path)
    assert logical_formatting(first_path) == logical_formatting(second_path)

    formatting = {
        title: (freeze_panes, auto_filter, formats)
        for title, freeze_panes, auto_filter, formats in logical_formatting(first_path)
    }
    assert formatting[VALID_SHEET][0] == "A2"
    assert formatting[INVALID_SHEET][0] == "A2"
    assert formatting[DUPLICATES_SHEET][0] == "A2"
    assert formatting[VALID_SHEET][1] is not None
    assert formatting[INVALID_SHEET][1] is not None
    assert formatting[DUPLICATES_SHEET][1] is not None
    assert formatting[VALID_SHEET][2][1][3] == "yyyy-mm-dd"


def test_round_trip_end_to_end_csv_and_xlsx(tmp_path: Path) -> None:
    input_path = tmp_path / "input"
    input_path.mkdir()
    (input_path / "orders.csv").write_text(
        "order_id,customer_name,email,order_date,amount,status,channel\n"
        "00123,Julia,,2026-08-20,10.00,paid,online\n"
        "duplicate,Ada,,2026-08-21,20.00,paid,store\n"
        "invalid,Grace,bad,2026-08-22,5.00,paid,online\n",
        encoding="utf-8",
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders"
    worksheet.append(
        [
            "order_id",
            "customer_name",
            "email",
            "order_date",
            "amount",
            "status",
            "region",
            "formula_note",
        ]
    )
    worksheet.append(
        ["duplicate", "Lin", None, date(2026, 8, 23), 30, "paid", "south", None]
    )
    worksheet.append(
        ["formula", "Eve", None, date(2026, 8, 24), 1, "pending", "north", "=1+1"]
    )
    workbook.save(input_path / "orders.xlsx")
    workbook.close()

    records = load_supported_files(discover_supported_files(input_path))
    result = process_records(records)
    summary = calculate_summary(result)
    output_path = tmp_path / "output" / "sales_report.xlsx"

    export_report(result, summary, output_path)

    output = load_workbook(output_path, data_only=False)
    try:
        assert tuple(output.sheetnames) == WORKSHEET_NAMES
        assert [output[SUMMARY_SHEET].cell(row=row, column=2).value for row in range(2, 7)] == [
            5,
            1,
            2,
            2,
            "10.00",
        ]
        assert output[VALID_SHEET]["A2"].value == "00123"
        assert output[VALID_SHEET]["A2"].data_type == "s"
        assert output[VALID_SHEET]["D2"].value == date(2026, 8, 20)
        assert output[VALID_SHEET]["E2"].value == "10.00"
        assert output[INVALID_SHEET].max_row == 3
        assert output[DUPLICATES_SHEET].max_row == 3

        invalid_headers = [cell.value for cell in output[INVALID_SHEET][1]]
        formula_column = invalid_headers.index("formula_note") + 1
        formula_cell = output[INVALID_SHEET].cell(row=3, column=formula_column)
        assert formula_cell.value == "'=1+1"
        assert formula_cell.data_type != "f"
        assert "formula_not_allowed" in output[INVALID_SHEET].cell(
            row=3,
            column=output[INVALID_SHEET].max_column,
        ).value
    finally:
        output.close()
