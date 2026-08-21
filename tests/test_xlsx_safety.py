from __future__ import annotations

import posixpath
import struct
import warnings
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZIP_STORED, BadZipFile, ZipFile, ZipInfo

import openpyxl.xml
import pytest
from defusedxml.common import DefusedXmlException
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

import src.xlsx_safety as safety_module
from src.processor import XlsxStructuralError, load_supported_files, load_xlsx_file
from src.xlsx_safety import (
    UnsafeXlsxPackageError,
    inspect_xlsx_package,
    validate_loaded_workbook_limits,
)


WORKSHEET_XML = """<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData><row r="1"><c r="A1"/></row></sheetData>
</worksheet>
"""
VALID_HEADER = [
    "order_id",
    "customer_name",
    "email",
    "order_date",
    "amount",
    "status",
]
VALID_ROW = ["001", "Julia", "julia@example.com", "2026-08-20", "1.00", "paid"]
CONTENT_TYPES_XML = """<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Override PartName="/xl/workbook.xml"
    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
</Types>
"""
WORKBOOK_XML = """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>
"""
WORKBOOK_WITH_SHEET_XML = """<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Orders" sheetId="1" r:id="rId1"/></sheets>
</workbook>
"""
WORKBOOK_RELS_XML = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>
"""


def write_package(
    path: Path,
    members: list[tuple[str, str | bytes]],
    *,
    compression: int = ZIP_DEFLATED,
) -> Path:
    with ZipFile(path, "w", compression=compression) as archive:
        for name, content in members:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                archive.writestr(name, content)
    return path


def write_minimal_xlsx_package(
    path: Path,
    members: list[tuple[str, str | bytes]] | None = None,
) -> Path:
    additional_members = members or []
    worksheet_names = [
        name for name, _content in additional_members if "worksheets/" in name
    ]
    relationships = "".join(
        '<Relationship Id="rId{index}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
        'relationships/worksheet" Target="/{name}"/>'.format(
            index=index,
            name=name,
        )
        for index, name in enumerate(worksheet_names, start=1)
    )
    sheets = "".join(
        '<sheet name="Sheet{index}" sheetId="{index}" r:id="rId{index}"/>'.format(
            index=index
        )
        for index, _name in enumerate(worksheet_names, start=1)
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/'
        f'relationships"><sheets>{sheets}</sheets></workbook>'
    )
    workbook_relationships = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
        f'relationships">{relationships}</Relationships>'
    )
    return write_package(
        path,
        [
            ("[Content_Types].xml", CONTENT_TYPES_XML),
            ("xl/workbook.xml", workbook_xml),
            ("xl/_rels/workbook.xml.rels", workbook_relationships),
            *additional_members,
        ],
    )


def save_workbook(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders"
    worksheet.append(VALID_HEADER)
    worksheet.append(VALID_ROW)
    workbook.save(path)
    workbook.close()
    return path


def rewrite_package_members(
    source: Path,
    target: Path,
    replacements: dict[str, bytes],
) -> Path:
    with ZipFile(source, "r") as archive:
        members = {
            member.filename: archive.read(member) for member in archive.infolist()
        }
    members.update(replacements)
    with ZipFile(target, "w", compression=ZIP_DEFLATED) as archive:
        for name, content in members.items():
            archive.writestr(name, content)
    return target


def relocate_workbook_package(
    source: Path,
    target: Path,
    *,
    workbook_part: str = "custom/book.xml",
    worksheet_folder: str = "custom/sheets",
) -> Path:
    with ZipFile(source, "r") as archive:
        members = {
            member.filename: archive.read(member) for member in archive.infolist()
        }

    old_workbook = "xl/workbook.xml"
    old_relationships = "xl/_rels/workbook.xml.rels"
    new_relationships = safety_module._workbook_relationships_path(workbook_part)
    members[workbook_part] = members.pop(old_workbook)
    relationships = members.pop(old_relationships)
    content_types = members["[Content_Types].xml"].replace(
        f"/{old_workbook}".encode(),
        f"/{workbook_part}".encode(),
    )
    root_relationships = members["_rels/.rels"].replace(
        old_workbook.encode(),
        workbook_part.encode(),
    )

    worksheet_names = sorted(
        name for name in members if name.startswith("xl/worksheets/")
    )
    for old_name in worksheet_names:
        new_name = f"{worksheet_folder}/{Path(old_name).name}"
        members[new_name] = members.pop(old_name)
        relationships = relationships.replace(
            f"/{old_name}".encode(),
            posixpath.relpath(
                new_name,
                posixpath.dirname(workbook_part),
            ).encode(),
        )
        content_types = content_types.replace(
            f"/{old_name}".encode(),
            f"/{new_name}".encode(),
        )

    members[new_relationships] = relationships
    members["[Content_Types].xml"] = content_types
    members["_rels/.rels"] = root_relationships
    with ZipFile(target, "w", compression=ZIP_DEFLATED) as archive:
        for name, content in members.items():
            archive.writestr(name, content)
    return target


def rewrite_stored_package(
    source: Path,
    target: Path,
    extra_members: list[tuple[str, bytes]] | None = None,
) -> Path:
    with ZipFile(source, "r") as archive:
        members = [(member.filename, archive.read(member)) for member in archive.infolist()]
    with ZipFile(target, "w", compression=ZIP_STORED) as archive:
        for name, content in [*members, *(extra_members or [])]:
            archive.writestr(name, content)
    return target


def patch_compression_method(path: Path, member_name: str, method: int) -> None:
    data = bytearray(path.read_bytes())
    encoded_name = member_name.encode()
    offset = 0
    patched_local = False
    patched_central = False
    while offset < len(data):
        signature = data[offset : offset + 4]
        if signature == b"PK\x03\x04":
            name_length, extra_length = struct.unpack_from("<HH", data, offset + 26)
            name = bytes(data[offset + 30 : offset + 30 + name_length])
            compressed_size = struct.unpack_from("<I", data, offset + 18)[0]
            if name == encoded_name:
                struct.pack_into("<H", data, offset + 8, method)
                patched_local = True
            offset += 30 + name_length + extra_length + compressed_size
        elif signature == b"PK\x01\x02":
            name_length, extra_length, comment_length = struct.unpack_from(
                "<HHH", data, offset + 28
            )
            name = bytes(data[offset + 46 : offset + 46 + name_length])
            if name == encoded_name:
                struct.pack_into("<H", data, offset + 10, method)
                patched_central = True
            offset += 46 + name_length + extra_length + comment_length
        else:
            offset += 1
    assert patched_local and patched_central
    path.write_bytes(data)


def corrupt_stored_member(path: Path, member_name: str) -> None:
    data = bytearray(path.read_bytes())
    encoded_name = member_name.encode()
    offset = 0
    while offset < len(data):
        if data[offset : offset + 4] != b"PK\x03\x04":
            offset += 1
            continue
        name_length, extra_length = struct.unpack_from("<HH", data, offset + 26)
        compressed_size = struct.unpack_from("<I", data, offset + 18)[0]
        name = bytes(data[offset + 30 : offset + 30 + name_length])
        payload_offset = offset + 30 + name_length + extra_length
        if name == encoded_name:
            assert compressed_size > 0
            data[payload_offset] ^= 0x01
            path.write_bytes(data)
            return
        offset = payload_offset + compressed_size
    pytest.fail(f"member not found: {member_name}")


def worksheet_xml(*, row: int = 1, column: str = "A") -> str:
    return WORKSHEET_XML.replace('row r="1"', f'row r="{row}"').replace(
        'c r="A1"',
        f'c r="{column}{row}"',
    )


def worksheet_with_sheet_data(sheet_data: str, *, extra_xml: str = "") -> str:
    return f'''<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
{extra_xml}<sheetData>{sheet_data}</sheetData>
</worksheet>'''


def test_requires_active_defusedxml_protection(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(tmp_path / "orders.xlsx", [("safe.txt", "safe")])
    monkeypatch.setattr(openpyxl.xml, "DEFUSEDXML", False)

    with pytest.raises(
        UnsafeXlsxPackageError,
        match="required defusedxml protection is not active",
    ):
        inspect_xlsx_package(source)


def test_inactive_defusedxml_stops_application_before_workbook_open(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(tmp_path / "orders.xlsx")
    monkeypatch.setattr(openpyxl.xml, "DEFUSEDXML", False)

    def fail_if_called(*_args: object, **_kwargs: object):
        pytest.fail("openpyxl must not load a workbook without XML protection")

    monkeypatch.setattr("src.processor.load_workbook", fail_if_called)

    with pytest.raises(XlsxStructuralError, match="defusedxml protection"):
        load_xlsx_file(source)


def test_current_environment_has_active_defusedxml_protection() -> None:
    assert openpyxl.xml.DEFUSEDXML is True


def test_rejects_minimal_entity_declaration_without_expansion(
    tmp_path: Path,
) -> None:
    xml = """<!DOCTYPE worksheet [<!ENTITY safe "not-expanded">]>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData><row r="1"><c r="A1"><v>&safe;</v></c></row></sheetData>
</worksheet>
"""
    source = write_minimal_xlsx_package(
        tmp_path / "entity.xlsx",
        [("xl/worksheets/sheet1.xml", xml)],
    )

    with pytest.raises(UnsafeXlsxPackageError) as captured:
        inspect_xlsx_package(source)

    assert isinstance(captured.value.__cause__, DefusedXmlException)


def test_prohibited_xml_is_controlled_structural_failure_without_partial_results(
    tmp_path: Path,
) -> None:
    xml = """<!DOCTYPE worksheet [<!ENTITY safe "not-expanded">]>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData><row r="1"><c r="A1"><v>&safe;</v></c></row></sheetData>
</worksheet>
"""
    valid = save_workbook(tmp_path / "a-valid.xlsx")
    invalid = write_minimal_xlsx_package(
        tmp_path / "b-invalid.xlsx",
        [("xl/worksheets/sheet1.xml", xml)],
    )

    with pytest.raises(XlsxStructuralError, match="package preflight failed"):
        load_supported_files([valid, invalid])


def test_rejects_invalid_zip(tmp_path: Path) -> None:
    source = tmp_path / "invalid.xlsx"
    source.write_bytes(b"not a ZIP archive")

    with pytest.raises(UnsafeXlsxPackageError, match="not a valid ZIP") as captured:
        inspect_xlsx_package(source)

    assert captured.value.__cause__ is not None


def test_rejects_empty_zip_as_controlled_structural_failure(tmp_path: Path) -> None:
    source = tmp_path / "empty.xlsx"
    with ZipFile(source, "w"):
        pass

    with pytest.raises(UnsafeXlsxPackageError, match="contains no members"):
        inspect_xlsx_package(source)

    with pytest.raises(XlsxStructuralError, match="package preflight failed"):
        load_xlsx_file(source)


def test_rejects_non_xlsx_zip(tmp_path: Path) -> None:
    source = write_package(tmp_path / "not-xlsx.xlsx", [("notes.txt", "hello")])

    with pytest.raises(UnsafeXlsxPackageError, match="Content_Types"):
        inspect_xlsx_package(source)

    with pytest.raises(XlsxStructuralError, match="package preflight failed"):
        load_xlsx_file(source)


def test_rejects_missing_declared_workbook_part(tmp_path: Path) -> None:
    source = write_package(
        tmp_path / "missing-workbook.xlsx",
        [("[Content_Types].xml", CONTENT_TYPES_XML)],
    )

    with pytest.raises(UnsafeXlsxPackageError, match="missing declared workbook"):
        inspect_xlsx_package(source)

    with pytest.raises(XlsxStructuralError, match="package preflight failed"):
        load_xlsx_file(source)


def test_rejects_missing_workbook_relationships(tmp_path: Path) -> None:
    source = write_package(
        tmp_path / "missing-relationships.xlsx",
        [
            ("[Content_Types].xml", CONTENT_TYPES_XML),
            ("xl/workbook.xml", WORKBOOK_XML),
        ],
    )

    with pytest.raises(UnsafeXlsxPackageError, match="workbook relationships"):
        inspect_xlsx_package(source)

    with pytest.raises(XlsxStructuralError, match="package preflight failed"):
        load_xlsx_file(source)


@pytest.mark.parametrize(
    "content_types",
    [
        CONTENT_TYPES_XML.replace("<Types ", "<NotTypes ").replace(
            "</Types>", "</NotTypes>"
        ),
        CONTENT_TYPES_XML.replace(
            "http://schemas.openxmlformats.org/package/2006/content-types",
            "urn:wrong-content-types",
        ),
    ],
)
def test_rejects_semantically_invalid_content_types_root(
    tmp_path: Path,
    content_types: str,
) -> None:
    source = write_package(
        tmp_path / "invalid-content-types.xlsx",
        [
            ("[Content_Types].xml", content_types),
            ("xl/workbook.xml", WORKBOOK_XML),
            ("xl/_rels/workbook.xml.rels", WORKBOOK_RELS_XML),
        ],
    )

    with pytest.raises(UnsafeXlsxPackageError, match="invalid root element"):
        inspect_xlsx_package(source)


@pytest.mark.parametrize(
    "content_types",
    [
        "<Types",
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>',
        CONTENT_TYPES_XML.replace(
            "</Types>",
            '''<Override PartName="/custom/book.xml"
ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
</Types>''',
        ),
    ],
)
def test_rejects_malformed_missing_or_duplicate_workbook_declaration(
    tmp_path: Path,
    content_types: str,
) -> None:
    source = write_package(
        tmp_path / "invalid-declaration.xlsx",
        [
            ("[Content_Types].xml", content_types),
            ("xl/workbook.xml", WORKBOOK_XML),
            ("xl/_rels/workbook.xml.rels", WORKBOOK_RELS_XML),
        ],
    )

    with pytest.raises(UnsafeXlsxPackageError):
        inspect_xlsx_package(source)


@pytest.mark.parametrize(
    "relationships",
    [
        "<Relationships",
        '<NotRelationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>',
        '<Relationships xmlns="urn:wrong-relationships"/>',
    ],
)
def test_rejects_malformed_or_semantically_invalid_relationships_root(
    tmp_path: Path,
    relationships: str,
) -> None:
    source = write_package(
        tmp_path / "invalid-relationships.xlsx",
        [
            ("[Content_Types].xml", CONTENT_TYPES_XML),
            ("xl/workbook.xml", WORKBOOK_WITH_SHEET_XML),
            ("xl/_rels/workbook.xml.rels", relationships),
        ],
    )

    with pytest.raises(UnsafeXlsxPackageError):
        inspect_xlsx_package(source)


@pytest.mark.parametrize(
    "workbook",
    [
        "<workbook",
        '<notWorkbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>',
        '<workbook xmlns="urn:wrong-workbook"/>',
    ],
)
def test_rejects_malformed_or_semantically_invalid_workbook_root(
    tmp_path: Path,
    workbook: str,
) -> None:
    source = write_package(
        tmp_path / "invalid-workbook.xlsx",
        [
            ("[Content_Types].xml", CONTENT_TYPES_XML),
            ("xl/workbook.xml", workbook),
            ("xl/_rels/workbook.xml.rels", WORKBOOK_RELS_XML),
        ],
    )

    with pytest.raises(UnsafeXlsxPackageError):
        inspect_xlsx_package(source)


def test_accepts_conventional_and_custom_workbook_relationship_paths(
    tmp_path: Path,
) -> None:
    conventional = save_workbook(tmp_path / "conventional.xlsx")
    custom = relocate_workbook_package(
        conventional,
        tmp_path / "custom.xlsx",
    )

    inspect_xlsx_package(conventional)
    inspect_xlsx_package(custom)
    records = load_xlsx_file(custom)

    assert records[0]["order_id"] == "001"
    assert records[0]["source_sheet"] == "Orders"


def test_accepts_multiple_custom_worksheets_with_relative_targets(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(VALID_HEADER)
    first.append(VALID_ROW)
    second = workbook.create_sheet("Second")
    second.append(VALID_HEADER)
    second.append(["002", "Ada", None, "2026-08-21", "2.00", "pending"])
    workbook.save(source)
    workbook.close()
    custom = relocate_workbook_package(source, tmp_path / "custom.xlsx")

    inspect_xlsx_package(custom)
    records = load_xlsx_file(custom)

    assert [record["source_sheet"] for record in records] == ["First", "Second"]


def test_chartsheet_does_not_bypass_or_count_as_worksheet(tmp_path: Path) -> None:
    source = tmp_path / "chartsheet.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["value"])
    worksheet.append([1])
    chart = BarChart()
    chart.add_data(Reference(worksheet, min_col=1, min_row=1, max_row=2))
    chartsheet = workbook.create_chartsheet("Chart")
    chartsheet.add_chart(chart)
    workbook.save(source)
    workbook.close()

    inspect_xlsx_package(source)


@pytest.mark.parametrize(
    ("limit_name", "limit", "message"),
    [
        ("MAX_ROWS_PER_WORKSHEET", 1, "row limit"),
        ("MAX_COLUMNS_PER_WORKSHEET", 5, "column limit"),
        ("MAX_TOTAL_LOGICAL_CELLS", 11, "total logical-cell limit"),
    ],
)
def test_custom_worksheet_limits_fail_before_openpyxl_load(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    limit_name: str,
    limit: int,
    message: str,
) -> None:
    source = save_workbook(tmp_path / "source.xlsx")
    custom = relocate_workbook_package(source, tmp_path / "custom.xlsx")
    monkeypatch.setattr(safety_module, limit_name, limit)

    def fail_if_called(*_args: object, **_kwargs: object) -> None:
        pytest.fail("openpyxl must not load a workbook rejected by preflight")

    monkeypatch.setattr("src.processor.load_workbook", fail_if_called)

    with pytest.raises(XlsxStructuralError, match=message):
        load_xlsx_file(custom)


@pytest.mark.parametrize(
    ("target", "target_mode", "message"),
    [
        ("../../../outside.xml", None, "escapes the package"),
        ("https://example.com/sheet.xml", None, "external or invalid target"),
        ("//example.com/sheet.xml", None, "external or invalid target"),
        ("http://[", None, "invalid target"),
        ("sheets\\sheet1.xml", None, "invalid target"),
        ("https://example.com/sheet.xml", "External", "external worksheet"),
        ("sheets/sheet1.xml", "external", "invalid TargetMode"),
        ("missing.xml", None, "targets a missing part"),
    ],
)
def test_rejects_invalid_worksheet_relationship_targets(
    tmp_path: Path,
    target: str,
    target_mode: str | None,
    message: str,
) -> None:
    mode = f' TargetMode="{target_mode}"' if target_mode else ""
    relationships = f'''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="{target}"{mode}/>
</Relationships>'''
    source = write_package(
        tmp_path / "invalid-target.xlsx",
        [
            ("[Content_Types].xml", CONTENT_TYPES_XML),
            ("xl/workbook.xml", WORKBOOK_WITH_SHEET_XML),
            ("xl/_rels/workbook.xml.rels", relationships),
        ],
    )

    with pytest.raises(UnsafeXlsxPackageError, match=message):
        inspect_xlsx_package(source)


def test_relationship_target_resolution_normalizes_safe_internal_segments() -> None:
    assert safety_module._resolve_relationship_target(
        "custom/deep/book.xml",
        "../sheets/./sheet1.xml",
    ) == "custom/sheets/sheet1.xml"
    assert safety_module._resolve_relationship_target(
        "custom/deep/book.xml",
        "/xl/worksheets/sheet1.xml",
    ) == "xl/worksheets/sheet1.xml"


@pytest.mark.parametrize("target", ["", "sheet\x00.xml", "../..", "/"])
def test_relationship_target_resolution_rejects_unsafe_paths(target: str) -> None:
    with pytest.raises(UnsafeXlsxPackageError):
        safety_module._resolve_relationship_target("custom/book.xml", target)


def test_rejects_duplicate_worksheet_targets(tmp_path: Path) -> None:
    relationships = '''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="/custom/sheet.xml"/>
<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="/custom/sheet.xml"/>
</Relationships>'''
    workbook = WORKBOOK_WITH_SHEET_XML.replace(
        "</sheets>",
        '<sheet name="Second" sheetId="2" r:id="rId2"/></sheets>',
    )
    source = write_package(
        tmp_path / "duplicate-target.xlsx",
        [
            ("[Content_Types].xml", CONTENT_TYPES_XML),
            ("xl/workbook.xml", workbook),
            ("xl/_rels/workbook.xml.rels", relationships),
            ("custom/sheet.xml", WORKSHEET_XML),
        ],
    )

    with pytest.raises(UnsafeXlsxPackageError, match="duplicate worksheet targets"):
        inspect_xlsx_package(source)


def test_rejects_sheet_relationship_with_non_worksheet_type(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(tmp_path / "source.xlsx")
    with ZipFile(source, "r") as archive:
        relationships = archive.read("xl/_rels/workbook.xml.rels").replace(
            b"/relationships/worksheet",
            b"/relationships/notWorksheet",
        )
    invalid = rewrite_package_members(
        source,
        tmp_path / "invalid-type.xlsx",
        {"xl/_rels/workbook.xml.rels": relationships},
    )

    def fail_if_called(*_args: object, **_kwargs: object) -> None:
        pytest.fail("openpyxl must not load invalid sheet relationships")

    monkeypatch.setattr("src.processor.load_workbook", fail_if_called)

    with pytest.raises(XlsxStructuralError, match="invalid type"):
        load_xlsx_file(invalid)


@pytest.mark.parametrize(
    "member_name",
    [
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/worksheets/sheet1.xml",
        "docProps/app.xml",
    ],
)
def test_rejects_unsupported_compression_before_member_read(
    tmp_path: Path,
    member_name: str,
) -> None:
    source = save_workbook(tmp_path / "orders.xlsx")
    patch_compression_method(source, member_name, 99)

    with pytest.raises(
        UnsafeXlsxPackageError,
        match="unsupported compression method 99",
    ):
        inspect_xlsx_package(source)

    with pytest.raises(XlsxStructuralError, match="package preflight failed"):
        load_xlsx_file(source)


def test_wraps_permission_error_opening_zip_and_preserves_cause(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(tmp_path / "orders.xlsx")
    permission_error = PermissionError("access denied after stat")

    def deny_open(*_args: object, **_kwargs: object) -> None:
        raise permission_error

    monkeypatch.setattr(safety_module, "ZipFile", deny_open)

    with pytest.raises(UnsafeXlsxPackageError, match="could not be opened") as captured:
        inspect_xlsx_package(source)

    assert captured.value.__cause__ is permission_error

    with pytest.raises(XlsxStructuralError) as structural:
        load_xlsx_file(source)

    assert structural.value.__cause__ is not None
    assert structural.value.__cause__.__cause__ is permission_error


def test_unexpected_runtime_error_is_not_masked(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(tmp_path / "orders.xlsx")
    programming_error = RuntimeError("unexpected bug")

    def fail_unexpectedly(*_args: object, **_kwargs: object) -> None:
        raise programming_error

    monkeypatch.setattr(safety_module, "_find_xlsx_workbook_part", fail_unexpectedly)

    with pytest.raises(RuntimeError) as captured:
        inspect_xlsx_package(source)

    assert captured.value is programming_error


@pytest.mark.parametrize(
    "member_name",
    [
        "xl/worksheets/sheet1.xml",
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/sharedStrings.xml",
        "xl/media/unused.bin",
    ],
)
def test_rejects_crc_corruption_in_every_member(
    tmp_path: Path,
    member_name: str,
) -> None:
    original = save_workbook(tmp_path / "original.xlsx")
    extra_members = [("xl/media/unused.bin", b"ignored but integrity checked")]
    if member_name == "xl/sharedStrings.xml":
        extra_members.append((member_name, b"<sst/>"))
    source = rewrite_stored_package(
        original,
        tmp_path / "corrupt.xlsx",
        extra_members,
    )
    corrupt_stored_member(source, member_name)

    with pytest.raises(XlsxStructuralError) as captured:
        load_xlsx_file(source)

    assert isinstance(captured.value.__cause__, UnsafeXlsxPackageError)
    assert isinstance(captured.value.__cause__.__cause__, BadZipFile)


def test_accepts_xlsx_at_exact_file_size_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = save_workbook(tmp_path / "orders.xlsx")
    monkeypatch.setattr("src.xlsx_safety.MAX_XLSX_FILE_SIZE", source.stat().st_size)

    inspect_xlsx_package(source)


def test_rejects_xlsx_above_file_size_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(tmp_path / "orders.xlsx", [("safe.txt", "safe")])
    monkeypatch.setattr(
        "src.xlsx_safety.MAX_XLSX_FILE_SIZE",
        source.stat().st_size - 1,
    )

    with pytest.raises(UnsafeXlsxPackageError, match="file exceeds the size limit"):
        inspect_xlsx_package(source)


def test_rejects_member_count_above_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [("one.txt", "1"), ("two.txt", "2")],
    )
    monkeypatch.setattr("src.xlsx_safety.MAX_ZIP_MEMBERS", 1)

    with pytest.raises(UnsafeXlsxPackageError, match="member limit of 1"):
        inspect_xlsx_package(source)


def test_accepts_exact_package_metadata_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [("one.txt", "12"), ("two.txt", "345")],
    )
    with ZipFile(source, "r") as archive:
        members = archive.infolist()
        total_compressed = sum(member.compress_size for member in members)
        total_uncompressed = sum(member.file_size for member in members)
        largest_member = max(member.file_size for member in members)

    monkeypatch.setattr("src.xlsx_safety.MAX_ZIP_MEMBERS", len(members))
    monkeypatch.setattr(
        "src.xlsx_safety.MAX_TOTAL_UNCOMPRESSED_SIZE",
        total_uncompressed,
    )
    monkeypatch.setattr(
        "src.xlsx_safety.MAX_INDIVIDUAL_MEMBER_SIZE",
        largest_member,
    )
    monkeypatch.setattr(
        "src.xlsx_safety.MAX_XLSX_FILE_SIZE",
        max(source.stat().st_size, total_compressed),
    )

    inspect_xlsx_package(source)


def test_rejects_individual_member_above_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(tmp_path / "orders.xlsx", [("large.txt", "12")])
    monkeypatch.setattr("src.xlsx_safety.MAX_INDIVIDUAL_MEMBER_SIZE", 1)

    with pytest.raises(UnsafeXlsxPackageError, match="individual size limit"):
        inspect_xlsx_package(source)


def test_rejects_total_uncompressed_size_above_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(
        tmp_path / "orders.xlsx",
        [("one.txt", "12"), ("two.txt", "34")],
    )
    monkeypatch.setattr("src.xlsx_safety.MAX_TOTAL_UNCOMPRESSED_SIZE", 3)

    with pytest.raises(UnsafeXlsxPackageError, match="total uncompressed-size"):
        inspect_xlsx_package(source)


def test_rejects_impossible_total_compressed_metadata(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(tmp_path / "orders.xlsx", [("safe.txt", "safe")])
    real_infolist = ZipFile.infolist

    def report_excessive_compressed_size(archive: ZipFile):
        members = real_infolist(archive)
        members[0].compress_size = source.stat().st_size + 1
        return members

    monkeypatch.setattr(ZipFile, "infolist", report_excessive_compressed_size)
    monkeypatch.setattr("src.xlsx_safety.MAX_XLSX_FILE_SIZE", source.stat().st_size)

    with pytest.raises(
        UnsafeXlsxPackageError,
        match="compressed member data exceeds",
    ):
        inspect_xlsx_package(source)


def test_rejects_material_individual_compression_ratio(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(tmp_path / "orders.xlsx", [("large.txt", "x" * 1_000)])
    monkeypatch.setattr("src.xlsx_safety.MIN_INDIVIDUAL_RATIO_SIZE", 10)
    monkeypatch.setattr("src.xlsx_safety.MAX_COMPRESSION_RATIO", 2)

    with pytest.raises(UnsafeXlsxPackageError, match="compression ratio limit"):
        inspect_xlsx_package(source)


def test_accepts_exact_compression_ratio_limit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("src.xlsx_safety.MIN_INDIVIDUAL_RATIO_SIZE", 0)
    monkeypatch.setattr("src.xlsx_safety.MAX_COMPRESSION_RATIO", 100)
    member = safety_module.ZipInfo("member")
    member.file_size = 100
    member.compress_size = 1

    assert not safety_module._member_ratio_exceeds_limit(member, 100)


def test_rejects_compression_ratio_immediately_above_limit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("src.xlsx_safety.MIN_INDIVIDUAL_RATIO_SIZE", 0)
    member = ZipInfo("member")
    member.file_size = 101
    member.compress_size = 1

    assert safety_module._member_ratio_exceeds_limit(member, 101)


def test_individual_ratio_materiality_uses_strict_greater_than() -> None:
    at_threshold = ZipInfo("at-threshold")
    at_threshold.file_size = safety_module.MIN_INDIVIDUAL_RATIO_SIZE
    at_threshold.compress_size = 1
    above_threshold = ZipInfo("above-threshold")
    above_threshold.file_size = safety_module.MIN_INDIVIDUAL_RATIO_SIZE + 1
    above_threshold.compress_size = 1

    assert not safety_module._member_ratio_exceeds_limit(
        at_threshold,
        safety_module.MIN_AGGREGATE_RATIO_SIZE,
    )
    assert safety_module._member_ratio_exceeds_limit(
        above_threshold,
        safety_module.MIN_AGGREGATE_RATIO_SIZE,
    )


def test_aggregate_ratio_materiality_uses_strict_greater_than() -> None:
    threshold = safety_module.MIN_AGGREGATE_RATIO_SIZE

    assert not safety_module._aggregate_ratio_exceeds_limit(threshold, 1)
    assert safety_module._aggregate_ratio_exceeds_limit(threshold + 1, 1)


def test_many_small_members_are_checked_when_only_aggregate_is_material() -> None:
    members = [ZipInfo(f"small-{index}") for index in range(1_000)]
    for member in members:
        member.file_size = 11_000
        member.compress_size = 100
    total_uncompressed = sum(member.file_size for member in members)
    total_compressed = sum(member.compress_size for member in members)

    assert all(
        member.file_size < safety_module.MIN_INDIVIDUAL_RATIO_SIZE
        for member in members
    )
    assert total_uncompressed > safety_module.MIN_AGGREGATE_RATIO_SIZE
    assert safety_module._aggregate_ratio_exceeds_limit(
        total_uncompressed,
        total_compressed,
    )


def test_rejects_material_aggregate_compression_ratio(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(
        tmp_path / "orders.xlsx",
        [("one.txt", "x" * 500), ("two.txt", "x" * 500)],
    )
    monkeypatch.setattr("src.xlsx_safety.MIN_INDIVIDUAL_RATIO_SIZE", 2_000)
    monkeypatch.setattr("src.xlsx_safety.MIN_AGGREGATE_RATIO_SIZE", 10)
    monkeypatch.setattr("src.xlsx_safety.MAX_COMPRESSION_RATIO", 2)

    with pytest.raises(UnsafeXlsxPackageError, match="aggregate compression ratio"):
        inspect_xlsx_package(source)


def test_zero_compressed_size_is_safe_only_for_empty_member(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [("empty.txt", b"")],
    )
    inspect_xlsx_package(source)

    real_infolist = ZipFile.infolist

    def report_impossible_zero_size(archive: ZipFile):
        members = real_infolist(archive)
        members[0].file_size = 2
        members[0].compress_size = 0
        return members

    monkeypatch.setattr(ZipFile, "infolist", report_impossible_zero_size)

    with pytest.raises(UnsafeXlsxPackageError, match="zero compressed size"):
        inspect_xlsx_package(source)


@pytest.mark.parametrize(
    ("name", "message"),
    [
        ("/absolute.xml", "absolute"),
        ("C:/absolute.xml", "absolute"),
        ("xl/../escape.xml", "parent traversal"),
        ("xl\\worksheets\\sheet1.xml", "backslash"),
    ],
)
def test_rejects_suspicious_member_names(
    tmp_path: Path,
    name: str,
    message: str,
) -> None:
    source = write_package(tmp_path / "orders.xlsx", [(name, "content")])

    with pytest.raises(UnsafeXlsxPackageError, match=message):
        inspect_xlsx_package(source)


def test_rejects_nul_in_original_member_name(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(tmp_path / "orders.xlsx", [("safe.txt", "safe")])
    real_infolist = ZipFile.infolist

    def report_nul_name(archive: ZipFile):
        members = real_infolist(archive)
        members[0].orig_filename = "nul\x00suffix.xml"
        return members

    monkeypatch.setattr(ZipFile, "infolist", report_nul_name)

    with pytest.raises(UnsafeXlsxPackageError, match="NUL"):
        inspect_xlsx_package(source)


def test_rejects_duplicate_member_names(tmp_path: Path) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [("duplicate.xml", "one"), ("duplicate.xml", "two")],
    )

    with pytest.raises(UnsafeXlsxPackageError, match="duplicate member name"):
        inspect_xlsx_package(source)


def test_rejects_encrypted_member_flag(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_package(tmp_path / "orders.xlsx", [("safe.txt", "safe")])
    real_infolist = ZipFile.infolist

    def report_encrypted_member(archive: ZipFile):
        members = real_infolist(archive)
        members[0].flag_bits |= 0x1
        return members

    monkeypatch.setattr(ZipFile, "infolist", report_encrypted_member)

    with pytest.raises(UnsafeXlsxPackageError, match="encrypted member"):
        inspect_xlsx_package(source)


def test_rejects_worksheet_count_above_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [
            ("xl/worksheets/sheet1.xml", WORKSHEET_XML),
            ("xl/worksheets/sheet2.xml", WORKSHEET_XML),
        ],
    )
    monkeypatch.setattr("src.xlsx_safety.MAX_WORKSHEETS", 1)

    with pytest.raises(UnsafeXlsxPackageError, match="worksheet limit of 1"):
        inspect_xlsx_package(source)


def test_accepts_exact_row_and_column_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [("xl/worksheets/sheet1.xml", worksheet_xml(row=2, column="B"))],
    )
    monkeypatch.setattr("src.xlsx_safety.MAX_ROWS_PER_WORKSHEET", 2)
    monkeypatch.setattr("src.xlsx_safety.MAX_COLUMNS_PER_WORKSHEET", 2)
    monkeypatch.setattr("src.xlsx_safety.MAX_TOTAL_LOGICAL_CELLS", 4)
    monkeypatch.setattr("src.xlsx_safety.MAX_WORKSHEETS", 1)

    inspect_xlsx_package(source)


@pytest.mark.parametrize(
    ("sheet_data", "limit_name", "limit", "message"),
    [
        ('<row r="2"/>', "MAX_ROWS_PER_WORKSHEET", 1, "row limit"),
        ('<row/><row/>', "MAX_ROWS_PER_WORKSHEET", 1, "row limit"),
        (
            '<row r="1"><c r="B1"/></row>',
            "MAX_COLUMNS_PER_WORKSHEET",
            1,
            "column limit",
        ),
        (
            '<row r="1"><c/><c/></row>',
            "MAX_COLUMNS_PER_WORKSHEET",
            1,
            "column limit",
        ),
        (
            '<row r="2"><c r="C2"/><c/><c r="B2"/><c/></row>',
            "MAX_TOTAL_LOGICAL_CELLS",
            7,
            "total logical-cell limit",
        ),
        (
            '<row r="3"><c/></row><row><c/></row>'
            '<row r="2"><c/></row><row><c/></row>',
            "MAX_TOTAL_LOGICAL_CELLS",
            3,
            "total logical-cell limit",
        ),
    ],
)
def test_counts_explicit_implicit_and_mixed_worksheet_references(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    sheet_data: str,
    limit_name: str,
    limit: int,
    message: str,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "references.xlsx",
        [
            (
                "xl/worksheets/sheet1.xml",
                worksheet_with_sheet_data(sheet_data),
            )
        ],
    )
    monkeypatch.setattr(safety_module, limit_name, limit)

    with pytest.raises(UnsafeXlsxPackageError, match=message):
        inspect_xlsx_package(source)


def test_accepts_implicit_references_exactly_at_all_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "implicit-exact.xlsx",
        [
            (
                "xl/worksheets/sheet1.xml",
                worksheet_with_sheet_data(
                    '<row><c/><c/></row><row><c/><c/></row>'
                ),
            )
        ],
    )
    monkeypatch.setattr(safety_module, "MAX_ROWS_PER_WORKSHEET", 2)
    monkeypatch.setattr(safety_module, "MAX_COLUMNS_PER_WORKSHEET", 2)
    monkeypatch.setattr(safety_module, "MAX_TOTAL_LOGICAL_CELLS", 4)

    inspect_xlsx_package(source)


@pytest.mark.parametrize(
    ("sheet_data", "row_limit", "column_limit", "logical_limit", "message"),
    [
        ('<row><c/></row><row><c/></row>', 1, 1, 2, "row limit"),
        ('<row><c/><c/></row>', 1, 1, 2, "column limit"),
        (
            '<row><c/><c/></row><row><c/><c/></row>',
            2,
            2,
            3,
            "logical-cell limit",
        ),
    ],
)
def test_implicit_reference_limit_plus_one_fails_before_openpyxl_load(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    sheet_data: str,
    row_limit: int,
    column_limit: int,
    logical_limit: int,
    message: str,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "implicit-limit.xlsx",
        [
            (
                "xl/worksheets/sheet1.xml",
                worksheet_with_sheet_data(sheet_data),
            )
        ],
    )
    monkeypatch.setattr(safety_module, "MAX_ROWS_PER_WORKSHEET", row_limit)
    monkeypatch.setattr(safety_module, "MAX_COLUMNS_PER_WORKSHEET", column_limit)
    monkeypatch.setattr(safety_module, "MAX_TOTAL_LOGICAL_CELLS", logical_limit)

    def fail_if_called(*_args: object, **_kwargs: object) -> None:
        pytest.fail("openpyxl must not load a worksheet above an implicit limit")

    monkeypatch.setattr("src.processor.load_workbook", fail_if_called)

    with pytest.raises(XlsxStructuralError, match=message):
        load_xlsx_file(source)


def test_ignores_row_and_cell_elements_outside_valid_sheet_data(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    extra_xml = '''
<ext:extension xmlns:ext="urn:example:extension">
  <ext:row r="999"><ext:c r="Z999"/></ext:row>
  <ext:c r="Z999"/><ext:row r="999"/>
</ext:extension>
<c r="Z999"/><row r="999"><c r="Z999"/></row>
'''
    sheet_data = '''
<row r="1"><c r="A1"/></row>
<ext:row xmlns:ext="urn:example:extension" r="999">
  <ext:c r="Z999"/>
</ext:row>
'''
    source = write_minimal_xlsx_package(
        tmp_path / "scoped-sheet-data.xlsx",
        [
            (
                "xl/worksheets/sheet1.xml",
                worksheet_with_sheet_data(sheet_data, extra_xml=extra_xml),
            )
        ],
    )
    monkeypatch.setattr(safety_module, "MAX_ROWS_PER_WORKSHEET", 1)
    monkeypatch.setattr(safety_module, "MAX_COLUMNS_PER_WORKSHEET", 1)
    monkeypatch.setattr(safety_module, "MAX_TOTAL_LOGICAL_CELLS", 1)

    inspect_xlsx_package(source)


@pytest.mark.parametrize(
    ("row", "column", "message"),
    [
        (3, "A", "row limit"),
        (1, "C", "column limit"),
    ],
)
def test_rejects_worksheet_dimension_above_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    row: int,
    column: str,
    message: str,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [("xl/worksheets/sheet1.xml", worksheet_xml(row=row, column=column))],
    )
    monkeypatch.setattr("src.xlsx_safety.MAX_ROWS_PER_WORKSHEET", 2)
    monkeypatch.setattr("src.xlsx_safety.MAX_COLUMNS_PER_WORKSHEET", 2)

    with pytest.raises(UnsafeXlsxPackageError, match=message):
        inspect_xlsx_package(source)


@pytest.mark.parametrize(
    ("reference_xml", "message"),
    [
        ('<row r="-1"><c r="A1"/></row>', "invalid row reference"),
        ('<row r="1"><c r=""/></row>', "empty cell reference"),
        ('<row r="1"><c r="A0"/></row>', "invalid cell reference"),
    ],
)
def test_rejects_non_positive_or_empty_worksheet_references_before_load(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    reference_xml: str,
    message: str,
) -> None:
    worksheet = f'''<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheetData>{reference_xml}</sheetData>
</worksheet>'''
    source = write_minimal_xlsx_package(
        tmp_path / "invalid-reference.xlsx",
        [("xl/worksheets/sheet1.xml", worksheet)],
    )

    def fail_if_called(*_args: object, **_kwargs: object) -> None:
        pytest.fail("openpyxl must not load invalid worksheet references")

    monkeypatch.setattr("src.processor.load_workbook", fail_if_called)

    with pytest.raises(XlsxStructuralError, match=message):
        load_xlsx_file(source)


def test_rejects_total_logical_cells_across_worksheets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [
            ("xl/worksheets/sheet1.xml", worksheet_xml(row=2, column="B")),
            ("xl/worksheets/sheet2.xml", worksheet_xml(row=2, column="B")),
        ],
    )
    monkeypatch.setattr("src.xlsx_safety.MAX_TOTAL_LOGICAL_CELLS", 7)

    with pytest.raises(UnsafeXlsxPackageError, match="total logical-cell limit"):
        inspect_xlsx_package(source)


def test_sparse_extreme_coordinate_exceeds_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [("xl/worksheets/sheet1.xml", worksheet_xml(row=100, column="Z"))],
    )
    monkeypatch.setattr("src.xlsx_safety.MAX_ROWS_PER_WORKSHEET", 100)
    monkeypatch.setattr("src.xlsx_safety.MAX_COLUMNS_PER_WORKSHEET", 26)
    monkeypatch.setattr("src.xlsx_safety.MAX_TOTAL_LOGICAL_CELLS", 2_599)

    with pytest.raises(UnsafeXlsxPackageError, match="total logical-cell limit"):
        inspect_xlsx_package(source)


def test_normal_xlsx_pipeline_remains_supported(tmp_path: Path) -> None:
    source = save_workbook(tmp_path / "orders.xlsx")

    records = load_xlsx_file(source)

    assert records[0]["order_id"] == "001"
    assert records[0]["source_sheet"] == "Orders"


def test_rechecks_limits_against_loaded_workbook_model(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=3, column=2, value="value")
    monkeypatch.setattr("src.xlsx_safety.MAX_ROWS_PER_WORKSHEET", 2)
    try:
        with pytest.raises(UnsafeXlsxPackageError, match="row limit"):
            validate_loaded_workbook_limits(workbook)
    finally:
        workbook.close()


def test_preflight_never_extracts_zip_members(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = write_minimal_xlsx_package(
        tmp_path / "orders.xlsx",
        [("safe.txt", "safe")],
    )

    def fail_extract(*_args: object, **_kwargs: object):
        pytest.fail("XLSX preflight must not extract ZIP members")

    monkeypatch.setattr(ZipFile, "extract", fail_extract)
    monkeypatch.setattr(ZipFile, "extractall", fail_extract)

    inspect_xlsx_package(source)
