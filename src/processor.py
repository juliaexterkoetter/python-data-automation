"""Input discovery and structural loading for sales data sources."""

from __future__ import annotations

import csv
import json
import logging
from collections import Counter
from collections.abc import Iterator, Sequence
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from typing import TypeAlias
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from src.validator import FormulaValue, ValidationError, normalize_and_validate_record
from src.xlsx_safety import (
    UnsafeXlsxPackageError,
    inspect_xlsx_package,
    validate_loaded_workbook_limits,
)


REQUIRED_COLUMNS = frozenset(
    {"order_id", "customer_name", "order_date", "amount", "status"}
)
RESERVED_TRACEABILITY_COLUMNS = frozenset(
    {"source_file", "source_sheet", "source_row"}
)
SUPPORTED_CSV_SUFFIX = ".csv"
SUPPORTED_XLSX_SUFFIX = ".xlsx"
SUPPORTED_INPUT_SUFFIXES = frozenset({SUPPORTED_CSV_SUFFIX, SUPPORTED_XLSX_SUFFIX})
CSV_MAX_FIELD_SIZE = 128 * 1024

LOGGER = logging.getLogger(__name__)

CsvValue: TypeAlias = str | None
LoadedRecord: TypeAlias = dict[str, object]


@dataclass(frozen=True)
class ProcessedRecord:
    """One normalized record with independent validation and duplicate state."""

    record: dict[str, object]
    validation_errors: tuple[ValidationError, ...]
    is_duplicate: bool

    @property
    def is_invalid(self) -> bool:
        return bool(self.validation_errors)

    @property
    def is_valid(self) -> bool:
        return not self.is_invalid and not self.is_duplicate


@dataclass(frozen=True)
class ProcessingResult:
    """Complete record accounting with intentionally overlapping classifications."""

    records: tuple[ProcessedRecord, ...]
    valid_records: tuple[ProcessedRecord, ...]
    invalid_records: tuple[ProcessedRecord, ...]
    duplicate_records: tuple[ProcessedRecord, ...]


class StructuralInputError(Exception):
    """Base exception for input failures that prevent a successful run."""


class InputDirectoryNotFoundError(StructuralInputError):
    """Raised when the configured input directory does not exist."""


class NoSupportedCsvFilesError(StructuralInputError):
    """Raised when the input directory contains no supported CSV files."""


class NoSupportedInputFilesError(StructuralInputError):
    """Raised when the input directory contains no supported input files."""


class InputDirectoryAccessError(StructuralInputError):
    """Raised when an operational filesystem error prevents input discovery."""

    def __init__(self, input_dir: Path, detail: str) -> None:
        self.input_dir = input_dir
        self.detail = detail
        super().__init__(f"Could not access input directory '{input_dir}': {detail}")


class CsvStructuralError(StructuralInputError):
    """Raised when a CSV source does not satisfy the structural contract."""

    def __init__(self, file_path: Path, detail: str) -> None:
        self.file_path = file_path
        self.detail = detail
        super().__init__(f"CSV structural error in '{file_path}': {detail}")


class XlsxStructuralError(StructuralInputError):
    """Raised when an XLSX source does not satisfy the structural contract."""

    def __init__(
        self,
        file_path: Path,
        detail: str,
        *,
        sheet_name: str | None = None,
    ) -> None:
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.detail = detail
        location = f"'{file_path}'"
        if sheet_name is not None:
            location += f", worksheet '{sheet_name}'"
        super().__init__(f"XLSX structural error in {location}: {detail}")


def _discover_files(input_dir: Path, suffixes: frozenset[str]) -> list[Path]:
    """Return matching regular non-symlink files in deterministic order."""
    try:
        if not input_dir.is_dir():
            raise InputDirectoryNotFoundError(
                f"Input directory does not exist: {input_dir}"
            )

        return sorted(
            (
                path
                for path in input_dir.iterdir()
                if not path.is_symlink()
                and path.is_file()
                and path.suffix.lower() in suffixes
            ),
            key=lambda path: (path.name.lower(), path.name),
        )
    except OSError as error:
        raise InputDirectoryAccessError(input_dir, str(error)) from error


def discover_csv_files(input_dir: Path) -> list[Path]:
    """Return supported CSV files directly inside *input_dir* in stable order."""
    csv_files = _discover_files(input_dir, frozenset({SUPPORTED_CSV_SUFFIX}))

    if not csv_files:
        raise NoSupportedCsvFilesError(
            f"No supported CSV files found in: {input_dir}"
        )

    return csv_files


def discover_supported_files(input_dir: Path) -> list[Path]:
    """Return supported CSV and XLSX files without recursion or symlink traversal."""
    source_files = _discover_files(input_dir, SUPPORTED_INPUT_SUFFIXES)
    if not source_files:
        raise NoSupportedInputFilesError(
            f"No supported input files found in: {input_dir}"
        )
    return source_files


def _detect_non_comma_delimiter(sample: str, file_path: Path) -> None:
    """Reject a detectable alternative delimiter without using it for loading."""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return

    if dialect.delimiter != ",":
        raise CsvStructuralError(
            file_path,
            f"expected comma delimiter, detected {dialect.delimiter!r}",
        )


@contextmanager
def _csv_field_size_policy() -> Iterator[None]:
    """Apply the deterministic CSV field limit and restore process state."""
    previous_limit = csv.field_size_limit()
    csv.field_size_limit(CSV_MAX_FIELD_SIZE)
    try:
        yield
    finally:
        csv.field_size_limit(previous_limit)


def _validate_header(header: list[str], file_path: Path) -> None:
    """Validate the exact, case-sensitive CSV header contract."""
    if not header or not any(header):
        raise CsvStructuralError(file_path, "missing header")

    if any(column == "" for column in header):
        raise CsvStructuralError(file_path, "empty header column")

    duplicate_columns = sorted(
        {column for column in header if header.count(column) > 1}
    )
    if duplicate_columns:
        raise CsvStructuralError(
            file_path,
            f"duplicate header columns: {', '.join(duplicate_columns)}",
        )

    reserved_columns = sorted(RESERVED_TRACEABILITY_COLUMNS.intersection(header))
    if reserved_columns:
        raise CsvStructuralError(
            file_path,
            f"reserved header columns: {', '.join(reserved_columns)}",
        )

    missing_columns = sorted(REQUIRED_COLUMNS.difference(header))
    if missing_columns:
        raise CsvStructuralError(
            file_path,
            f"missing required columns: {', '.join(missing_columns)}",
        )


def load_csv_file(file_path: Path) -> list[LoadedRecord]:
    """Validate and load one CSV file with source traceability metadata."""
    try:
        with _csv_field_size_policy(), file_path.open(
            encoding="utf-8-sig",
            errors="strict",
            newline="",
        ) as file:
            sample = file.read(8192)
            if not sample:
                raise CsvStructuralError(file_path, "missing header")

            _detect_non_comma_delimiter(sample, file_path)
            file.seek(0)
            reader = csv.reader(file, delimiter=",", strict=True)

            try:
                header = next(reader)
            except StopIteration as error:
                raise CsvStructuralError(file_path, "missing header") from error

            _validate_header(header, file_path)
            records: list[LoadedRecord] = []

            while True:
                source_row = reader.line_num + 1
                try:
                    row = next(reader)
                except StopIteration:
                    break

                if len(row) > len(header):
                    raise CsvStructuralError(
                        file_path,
                        f"row {source_row} contains more values than the header",
                    )

                values: list[CsvValue] = [*row, *([None] * (len(header) - len(row)))]
                record: LoadedRecord = dict(zip(header, values, strict=True))
                record["source_file"] = file_path.name
                record["source_sheet"] = None
                record["source_row"] = source_row
                records.append(record)

            if not records:
                raise CsvStructuralError(file_path, "file contains no data records")

            return records
    except UnicodeDecodeError as error:
        raise CsvStructuralError(file_path, "file is not valid UTF-8") from error
    except csv.Error as error:
        if "field larger than field limit" in str(error):
            detail = (
                f"field exceeds the maximum size of "
                f"{CSV_MAX_FIELD_SIZE} characters"
            )
        else:
            detail = f"malformed CSV: {error}"
        raise CsvStructuralError(file_path, detail) from error
    except OSError as error:
        raise CsvStructuralError(file_path, f"file could not be read: {error}") from error


def load_csv_files(file_paths: Sequence[Path]) -> list[LoadedRecord]:
    """Load all CSV sources, failing the complete operation on any source error."""
    records: list[LoadedRecord] = []
    for file_path in file_paths:
        records.extend(load_csv_file(file_path))
    return records


def _last_effective_data_row(worksheet: object) -> int | None:
    """Return the final row containing a value below the header."""
    last_row: int | None = None
    for row in worksheet.iter_rows(min_row=2):
        if any(cell.value is not None for cell in row):
            last_row = row[0].row
    return last_row


def _xlsx_header(worksheet: object, file_path: Path) -> tuple[list[object], int]:
    """Read and structurally validate the physical row-1 XLSX header."""
    sheet_name = worksheet.title
    header_cells = list(next(worksheet.iter_rows(min_row=1, max_row=1)))

    if any(cell.data_type == "f" for cell in header_cells):
        raise XlsxStructuralError(
            file_path,
            "formula in header row",
            sheet_name=sheet_name,
        )

    last_used_column = 0
    for cell in header_cells:
        if cell.value is not None:
            last_used_column = max(last_used_column, cell.column)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.value is not None:
                last_used_column = max(last_used_column, cell.column)

    if last_used_column == 0:
        return [], 0

    header = [cell.value for cell in header_cells[:last_used_column]]
    for column_index, value in enumerate(header, start=1):
        if value is None:
            raise XlsxStructuralError(
                file_path,
                f"header column {column_index} is empty before the last used column",
                sheet_name=sheet_name,
            )
        if not isinstance(value, str):
            raise XlsxStructuralError(
                file_path,
                f"header column {column_index} must be text",
                sheet_name=sheet_name,
            )

    named_header = [value for value in header if isinstance(value, str)]
    duplicate_columns = sorted(
        {column for column in named_header if named_header.count(column) > 1}
    )
    if duplicate_columns:
        raise XlsxStructuralError(
            file_path,
            f"duplicate header columns: {', '.join(duplicate_columns)}",
            sheet_name=sheet_name,
        )

    reserved_columns = sorted(
        RESERVED_TRACEABILITY_COLUMNS.intersection(named_header)
    )
    if reserved_columns:
        raise XlsxStructuralError(
            file_path,
            f"reserved header columns: {', '.join(reserved_columns)}",
            sheet_name=sheet_name,
        )

    return header, last_used_column


def _formula_attribute(value: object) -> str:
    """Serialize one documented formula attribute without object repr fallbacks."""
    if not isinstance(value, (str, bool, type(None))):
        raise TypeError(f"unsupported formula attribute type: {type(value).__name__}")
    return json.dumps(value, ensure_ascii=True, separators=(",", ":"))


def _retain_formula(
    value: object,
    file_path: Path,
    sheet_name: str,
    coordinate: str,
) -> FormulaValue:
    """Return a deterministic, non-executable trace of an XLSX formula."""
    if isinstance(value, str):
        return FormulaValue(value)

    if isinstance(value, ArrayFormula):
        if isinstance(value.text, str):
            return FormulaValue(value.text)
        try:
            reference = _formula_attribute(value.ref)
        except TypeError as error:
            raise XlsxStructuralError(
                file_path,
                f"unsupported array formula at cell {coordinate}: {error}",
                sheet_name=sheet_name,
            ) from error
        return FormulaValue(f"array(ref={reference}, text=null)")

    if isinstance(value, DataTableFormula):
        attributes = (
            ("ref", value.ref),
            ("ca", value.ca),
            ("dt2D", value.dt2D),
            ("dtr", value.dtr),
            ("r1", value.r1),
            ("r2", value.r2),
            ("del1", value.del1),
            ("del2", value.del2),
        )
        try:
            serialized = ", ".join(
                f"{name}={_formula_attribute(attribute)}"
                for name, attribute in attributes
            )
        except TypeError as error:
            raise XlsxStructuralError(
                file_path,
                f"unsupported data table formula at cell {coordinate}: {error}",
                sheet_name=sheet_name,
            ) from error
        return FormulaValue(f"dataTable({serialized})")

    raise XlsxStructuralError(
        file_path,
        (
            f"unsupported formula representation at cell {coordinate}: "
            f"{type(value).__name__}"
        ),
        sheet_name=sheet_name,
    )


def load_xlsx_file(file_path: Path) -> list[LoadedRecord]:
    """Validate and load all usable worksheets from one XLSX workbook."""
    try:
        inspect_xlsx_package(file_path)
    except UnsafeXlsxPackageError as error:
        raise XlsxStructuralError(
            file_path,
            f"package preflight failed: {error}",
        ) from error

    try:
        workbook = load_workbook(
            file_path,
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except (
        BadZipFile,
        DefusedXmlException,
        InvalidFileException,
        OSError,
        ParseError,
        ValueError,
    ) as error:
        raise XlsxStructuralError(
            file_path,
            f"workbook could not be read: {error}",
        ) from error

    try:
        try:
            validate_loaded_workbook_limits(workbook)
        except UnsafeXlsxPackageError as error:
            raise XlsxStructuralError(
                file_path,
                f"workbook limit validation failed: {error}",
            ) from error

        records: list[LoadedRecord] = []
        usable_worksheets = 0

        for worksheet in workbook.worksheets:
            header, last_used_column = _xlsx_header(worksheet, file_path)
            named_header = {value for value in header if isinstance(value, str)}
            required_present = REQUIRED_COLUMNS.intersection(named_header)

            if not required_present:
                LOGGER.info(
                    "Skipped auxiliary worksheet '%s' in '%s' as non-data.",
                    worksheet.title,
                    file_path,
                )
                continue

            missing_columns = sorted(REQUIRED_COLUMNS.difference(named_header))
            if missing_columns:
                raise XlsxStructuralError(
                    file_path,
                    f"missing required columns: {', '.join(missing_columns)}",
                    sheet_name=worksheet.title,
                )

            last_data_row = _last_effective_data_row(worksheet)
            if last_data_row is None:
                LOGGER.info(
                    "Skipped header-only worksheet '%s' in '%s' as unusable data.",
                    worksheet.title,
                    file_path,
                )
                continue

            usable_worksheets += 1
            column_names = {
                index: value
                for index, value in enumerate(header, start=1)
                if isinstance(value, str)
            }
            for row_index in range(2, last_data_row + 1):
                record: LoadedRecord = {}
                for column_index, column_name in column_names.items():
                    cell = worksheet.cell(row=row_index, column=column_index)
                    value: object = cell.value
                    if cell.data_type == "f":
                        value = _retain_formula(
                            value,
                            file_path,
                            worksheet.title,
                            cell.coordinate,
                        )
                    record[column_name] = value

                record["source_file"] = file_path.name
                record["source_sheet"] = worksheet.title
                record["source_row"] = row_index
                records.append(record)

        if usable_worksheets == 0:
            raise XlsxStructuralError(file_path, "workbook contains no usable worksheet")

        return records
    except XlsxStructuralError:
        raise
    except (
        BadZipFile,
        DefusedXmlException,
        InvalidFileException,
        OSError,
        ParseError,
    ) as error:
        raise XlsxStructuralError(
            file_path,
            f"workbook could not be read: {error}",
        ) from error
    finally:
        workbook.close()


def load_supported_files(file_paths: Sequence[Path]) -> list[LoadedRecord]:
    """Load every supported source, failing without returning partial results."""
    records: list[LoadedRecord] = []
    for file_path in file_paths:
        suffix = file_path.suffix.lower()
        if suffix == SUPPORTED_CSV_SUFFIX:
            records.extend(load_csv_file(file_path))
        elif suffix == SUPPORTED_XLSX_SUFFIX:
            records.extend(load_xlsx_file(file_path))
        else:
            raise StructuralInputError(f"Unsupported input file: {file_path}")
    return records


def process_records(records: Sequence[LoadedRecord]) -> ProcessingResult:
    """Normalize, validate, deduplicate, and classify loaded records."""
    validated_records = [normalize_and_validate_record(record) for record in records]
    id_counts = Counter(
        order_id
        for result in validated_records
        if isinstance((order_id := result.record.get("order_id")), str)
        and order_id
    )
    duplicate_ids = {order_id for order_id, count in id_counts.items() if count > 1}

    processed_records = tuple(
        ProcessedRecord(
            record=result.record,
            validation_errors=result.errors,
            is_duplicate=result.record.get("order_id") in duplicate_ids,
        )
        for result in validated_records
    )

    return ProcessingResult(
        records=processed_records,
        valid_records=tuple(record for record in processed_records if record.is_valid),
        invalid_records=tuple(
            record for record in processed_records if record.is_invalid
        ),
        duplicate_records=tuple(
            record for record in processed_records if record.is_duplicate
        ),
    )
