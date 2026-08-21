"""Deterministic and failure-safe Excel report generation."""

from __future__ import annotations

import math
import os
import sys
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Final, Sequence
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from src.processor import ProcessedRecord, ProcessingResult
from src.summary import ProcessingSummary
from src.validator import FormulaValue, ValidationError


SUMMARY_SHEET: Final = "Summary"
VALID_SHEET: Final = "Valid Records"
INVALID_SHEET: Final = "Invalid Records"
DUPLICATES_SHEET: Final = "Duplicates"
WORKSHEET_NAMES: Final = (
    SUMMARY_SHEET,
    VALID_SHEET,
    INVALID_SHEET,
    DUPLICATES_SHEET,
)

KNOWN_COLUMNS: Final = (
    "order_id",
    "customer_name",
    "email",
    "order_date",
    "amount",
    "status",
)
TRACEABILITY_COLUMNS: Final = ("source_file", "source_sheet", "source_row")
VALIDATION_ERRORS_COLUMN: Final = "validation_errors"
ASCII_LEADING_WHITESPACE: Final = " \t\r\n"
FORMULA_TRIGGERS: Final = frozenset("=+-@")
XLSX_MAX_ROWS: Final = 1_048_576
XLSX_MAX_COLUMNS: Final = 16_384
XLSX_MAX_CELL_CHARACTERS: Final = 32_767
MAX_EXACT_EXCEL_INTEGER: Final = 2**53

BOLD_FONT: Final = Font(bold=True)
DATE_FORMAT: Final = "yyyy-mm-dd"
DATETIME_FORMAT: Final = "yyyy-mm-dd hh:mm:ss"
TIME_FORMAT: Final = "hh:mm:ss"
TEXT_FORMAT: Final = "@"
GENERAL_FORMAT: Final = "General"


class ReportExportError(Exception):
    """A predictable export failure that prevents report publication."""

    def __init__(self, output_path: Path, detail: str) -> None:
        self.output_path = output_path
        self.detail = detail
        super().__init__(f"Report export failed for '{output_path}': {detail}")


@dataclass(frozen=True)
class _LogicalCell:
    value: object
    data_type: str | None
    number_format: str
    bold: bool
    input_controlled: bool = field(compare=False)


@dataclass(frozen=True)
class _LogicalWorksheet:
    title: str
    cells: tuple[tuple[_LogicalCell, ...], ...]
    freeze_panes: str | None
    auto_filter: str | None


LogicalWorkbook = tuple[_LogicalWorksheet, ...]


def _raise_data_error(output_path: Path, detail: str) -> None:
    error = ValueError(detail)
    raise ReportExportError(output_path, detail) from error


def _contains_invalid_xml_character(value: str) -> bool:
    for character in value:
        codepoint = ord(character)
        if codepoint in (0x09, 0x0A, 0x0D):
            continue
        if 0x20 <= codepoint <= 0xD7FF:
            continue
        if 0xE000 <= codepoint <= 0xFFFD:
            continue
        if 0x10000 <= codepoint <= 0x10FFFF:
            continue
        return True
    return False


def _text_length(value: str, output_path: Path) -> int:
    try:
        return len(value.encode("utf-16-le")) // 2
    except UnicodeEncodeError as error:
        raise ReportExportError(
            output_path,
            "text contains a character that cannot be encoded in XLSX",
        ) from error


def _validate_text(value: str, output_path: Path) -> None:
    if _contains_invalid_xml_character(value):
        _raise_data_error(
            output_path,
            "text contains an XML-incompatible character",
        )
    if _text_length(value, output_path) > XLSX_MAX_CELL_CHARACTERS:
        _raise_data_error(
            output_path,
            f"cell text exceeds the XLSX limit of {XLSX_MAX_CELL_CHARACTERS} characters",
        )


def _is_formula_triggering_text(value: str) -> bool:
    significant = value.lstrip(ASCII_LEADING_WHITESPACE)
    return bool(significant and significant[0] in FORMULA_TRIGGERS)


def _protect_text(value: str, output_path: Path) -> str:
    protected = (
        f"'{value}"
        if _is_formula_triggering_text(value)
        else value
    )
    _validate_text(protected, output_path)
    return protected


def _money_text(value: Decimal, output_path: Path) -> str:
    if not value.is_finite() or value.as_tuple().exponent != -2:
        _raise_data_error(
            output_path,
            "monetary values must be finite Decimal instances normalized to two places",
        )
    return format(value, ".2f")


def _independent_expected_money_text(
    value: object,
    output_path: Path,
) -> str:
    if not isinstance(value, Decimal):
        _raise_data_error(output_path, "monetary invariant requires a Decimal value")
    if not value.is_finite() or value.as_tuple().exponent != -2:
        _raise_data_error(
            output_path,
            "monetary invariant requires a finite Decimal normalized to two places",
        )

    decimal_tuple = value.as_tuple()
    digits = "".join(str(digit) for digit in decimal_tuple.digits).rjust(3, "0")
    sign = "-" if decimal_tuple.sign else ""
    expected = f"{sign}{digits[:-2]}.{digits[-2:]}"
    integer_part, fractional_part = expected.lstrip("-").split(".", maxsplit=1)
    if not integer_part or len(fractional_part) != 2:
        _raise_data_error(output_path, "monetary invariant produced an invalid value")
    return expected


def _logical_cell(
    value: object,
    output_path: Path,
    *,
    column_name: str | None = None,
    input_controlled: bool = True,
    bold: bool = False,
) -> _LogicalCell:
    if value is None:
        return _LogicalCell(None, None, GENERAL_FORMAT, bold, input_controlled)

    if isinstance(value, FormulaValue):
        if not isinstance(value.expression, str):
            _raise_data_error(output_path, "formula expressions must be text")
        protected = _protect_text(value.expression, output_path)
        return _LogicalCell(
            None if protected == "" else protected,
            None if protected == "" else "s",
            TEXT_FORMAT,
            bold,
            input_controlled,
        )

    if isinstance(value, str):
        protected = (
            _protect_text(value, output_path)
            if input_controlled
            else value
        )
        if not input_controlled:
            _validate_text(value, output_path)
        return _LogicalCell(
            None if protected == "" else protected,
            None if protected == "" else "s",
            TEXT_FORMAT,
            bold,
            input_controlled,
        )

    if isinstance(value, Decimal):
        if not value.is_finite():
            _raise_data_error(output_path, "non-finite Decimal values are not supported")
        text_value = (
            _money_text(value, output_path)
            if column_name == "amount"
            else format(value, "f")
        )
        protected = (
            _protect_text(text_value, output_path)
            if input_controlled
            else text_value
        )
        return _LogicalCell(protected, "s", TEXT_FORMAT, bold, input_controlled)

    if isinstance(value, datetime):
        if value.tzinfo is not None:
            _raise_data_error(output_path, "timezone-aware datetimes are not supported")
        if value.microsecond % 1000:
            _raise_data_error(
                output_path,
                "datetime precision finer than one millisecond is not supported by XLSX",
            )
        return _LogicalCell(value, "d", DATETIME_FORMAT, bold, input_controlled)

    if isinstance(value, date):
        return _LogicalCell(value, "d", DATE_FORMAT, bold, input_controlled)

    if isinstance(value, time):
        if value.tzinfo is not None:
            _raise_data_error(output_path, "timezone-aware times are not supported")
        if value.microsecond % 1000:
            _raise_data_error(
                output_path,
                "time precision finer than one millisecond is not supported by XLSX",
            )
        return _LogicalCell(value, "d", TIME_FORMAT, bold, input_controlled)

    if isinstance(value, bool):
        return _LogicalCell(value, "b", GENERAL_FORMAT, bold, input_controlled)

    if isinstance(value, int):
        if abs(value) > MAX_EXACT_EXCEL_INTEGER:
            text_value = format(value, "d")
            protected = (
                _protect_text(text_value, output_path)
                if input_controlled
                else text_value
            )
            return _LogicalCell(
                protected,
                "s",
                TEXT_FORMAT,
                bold,
                input_controlled,
            )
        return _LogicalCell(value, "n", GENERAL_FORMAT, bold, input_controlled)

    if isinstance(value, float):
        if not math.isfinite(value):
            _raise_data_error(output_path, "non-finite float values are not supported")
        if value == 0.0 and math.copysign(1.0, value) < 0:
            _raise_data_error(
                output_path,
                "negative signed float zero cannot survive XLSX serialization",
            )
        serialized_value = format(value, ".16g")
        if float(serialized_value) != value or (
            value == 0.0
            and math.copysign(1.0, value) != math.copysign(
                1.0,
                float(serialized_value),
            )
        ):
            _raise_data_error(
                output_path,
                "float value cannot survive exact XLSX numeric serialization",
            )
        return _LogicalCell(value, "n", GENERAL_FORMAT, bold, input_controlled)

    _raise_data_error(
        output_path,
        f"unsupported cell value type: {type(value).__name__}",
    )


def _validation_errors_text(
    errors: Sequence[ValidationError],
    output_path: Path,
) -> str:
    serialized_errors: list[str] = []
    for error in errors:
        if not isinstance(error, ValidationError) or not all(
            isinstance(value, str)
            for value in (error.field, error.code, error.message)
        ):
            _raise_data_error(
                output_path,
                "validation errors must contain textual field, code, and message values",
            )
        serialized_errors.append(
            f"{error.field} [{error.code}]: {error.message}"
        )
    return " | ".join(serialized_errors)


def _validate_projection_integrity(
    result: ProcessingResult,
    output_path: Path,
) -> None:
    if any(not isinstance(record, ProcessedRecord) for record in result.records):
        _raise_data_error(
            output_path,
            "processing result records must contain ProcessedRecord instances",
        )
    if any(not isinstance(record.record, dict) for record in result.records):
        _raise_data_error(
            output_path,
            "processed record data must use dictionaries",
        )
    record_ids = [id(record) for record in result.records]
    if len(record_ids) != len(set(record_ids)):
        _raise_data_error(
            output_path,
            "processing result contains the same record instance more than once",
        )

    expected_ids = set(record_ids)
    covered_ids: set[int] = set()
    for projection_name, projection in (
        ("valid_records", result.valid_records),
        ("invalid_records", result.invalid_records),
        ("duplicate_records", result.duplicate_records),
    ):
        if any(not isinstance(record, ProcessedRecord) for record in projection):
            _raise_data_error(
                output_path,
                f"{projection_name} must contain ProcessedRecord instances",
            )
        projection_ids = [id(record) for record in projection]
        if len(projection_ids) != len(set(projection_ids)):
            _raise_data_error(
                output_path,
                f"{projection_name} repeats a record instance",
            )
        if not set(projection_ids).issubset(expected_ids):
            _raise_data_error(
                output_path,
                f"{projection_name} contains a record outside result.records",
            )
        covered_ids.update(projection_ids)

    if covered_ids != expected_ids:
        _raise_data_error(
            output_path,
            "one or more processed records are absent from every output projection",
        )


def _record_columns(
    result: ProcessingResult,
    output_path: Path,
) -> tuple[tuple[str, ...], str]:
    fixed_columns = frozenset((*KNOWN_COLUMNS, *TRACEABILITY_COLUMNS))
    for processed_record in result.records:
        for column in processed_record.record:
            if not isinstance(column, str):
                _raise_data_error(
                    output_path,
                    "processed record column names must be text",
                )
            if column == "":
                _raise_data_error(
                    output_path,
                    "processed record column names must not be empty",
                )
    extra_columns = sorted(
        {
            column
            for processed_record in result.records
            for column in processed_record.record
            if column not in fixed_columns
        }
    )
    base_columns = (*KNOWN_COLUMNS, *extra_columns, *TRACEABILITY_COLUMNS)

    error_column = VALIDATION_ERRORS_COLUMN
    suffix = 2
    while error_column in base_columns:
        error_column = f"{VALIDATION_ERRORS_COLUMN}_{suffix}"
        suffix += 1

    return base_columns, error_column


def _validate_worksheet_size(
    row_count: int,
    column_count: int,
    output_path: Path,
    sheet_name: str,
) -> None:
    if row_count > XLSX_MAX_ROWS:
        _raise_data_error(
            output_path,
            f"worksheet '{sheet_name}' exceeds the XLSX row limit of {XLSX_MAX_ROWS}",
        )
    if column_count > XLSX_MAX_COLUMNS:
        _raise_data_error(
            output_path,
            f"worksheet '{sheet_name}' exceeds the XLSX column limit of {XLSX_MAX_COLUMNS}",
        )


def _header_cells(
    columns: Sequence[str],
    input_columns: frozenset[str],
    output_path: Path,
) -> tuple[_LogicalCell, ...]:
    rendered_columns = [
        _protect_text(column_name, output_path)
        if column_name in input_columns
        else column_name
        for column_name in columns
    ]
    for rendered_column in rendered_columns:
        _validate_text(rendered_column, output_path)
    if len(rendered_columns) != len(set(rendered_columns)):
        _raise_data_error(
            output_path,
            "input column names collide after formula-injection protection",
        )

    return tuple(
        _logical_cell(
            column_name,
            output_path,
            input_controlled=original_name in input_columns,
            bold=True,
        )
        for original_name, column_name in zip(
            columns,
            rendered_columns,
            strict=True,
        )
    )


def _record_worksheet_model(
    sheet_name: str,
    records: Sequence[ProcessedRecord],
    base_columns: tuple[str, ...],
    error_column: str,
    output_path: Path,
    *,
    include_errors: bool,
) -> _LogicalWorksheet:
    columns = (*base_columns, error_column) if include_errors else base_columns
    _validate_worksheet_size(
        len(records) + 1,
        len(columns),
        output_path,
        sheet_name,
    )
    generated_columns = frozenset(
        (*KNOWN_COLUMNS, *TRACEABILITY_COLUMNS, error_column)
    )
    input_columns = frozenset(
        column for column in columns if column not in generated_columns
    )
    rows: list[tuple[_LogicalCell, ...]] = [
        _header_cells(columns, input_columns, output_path)
    ]

    for processed_record in records:
        row = [
            _logical_cell(
                processed_record.record.get(column_name),
                output_path,
                column_name=column_name,
            )
            for column_name in base_columns
        ]

        if include_errors:
            row.append(
                _logical_cell(
                _validation_errors_text(
                    processed_record.validation_errors,
                    output_path,
                ),
                output_path,
                )
            )
        rows.append(tuple(row))

    last_column = get_column_letter(len(columns))
    return _LogicalWorksheet(
        title=sheet_name,
        cells=tuple(rows),
        freeze_panes="A2",
        auto_filter=f"A1:{last_column}{len(rows)}",
    )


def _summary_model(
    summary: ProcessingSummary,
    output_path: Path,
) -> _LogicalWorksheet:
    for field_name in (
        "total_records",
        "valid_records",
        "invalid_records",
        "duplicate_records",
    ):
        value = getattr(summary, field_name)
        if type(value) is not int or value < 0:
            _raise_data_error(
                output_path,
                f"summary {field_name} must be a non-negative integer",
            )
    if not isinstance(summary.total_paid_amount, Decimal):
        _raise_data_error(
            output_path,
            "summary total_paid_amount must be Decimal",
        )

    rows: tuple[tuple[str, object], ...] = (
        ("Total Records", summary.total_records),
        ("Valid Records", summary.valid_records),
        ("Invalid Records", summary.invalid_records),
        ("Duplicate Records", summary.duplicate_records),
        ("Total Paid Amount (USD)", summary.total_paid_amount),
    )
    _validate_worksheet_size(8, 2, output_path, SUMMARY_SHEET)

    logical_rows: list[tuple[_LogicalCell, ...]] = [
        (
            _logical_cell(
                "Metric",
                output_path,
                input_controlled=False,
                bold=True,
            ),
            _logical_cell(
                "Value",
                output_path,
                input_controlled=False,
                bold=True,
            ),
        )
    ]

    for label, value in rows:
        if isinstance(value, Decimal):
            value = _money_text(value, output_path)
        logical_rows.append(
            (
                _logical_cell(label, output_path, input_controlled=False),
                _logical_cell(value, output_path, input_controlled=False),
            )
        )

    empty_cell = _logical_cell(None, output_path, input_controlled=False)
    logical_rows.append((empty_cell, empty_cell))
    logical_rows.append(
        (
            _logical_cell(
                "Invalid and duplicate record counts may overlap.",
                output_path,
                input_controlled=False,
            ),
            empty_cell,
        )
    )
    return _LogicalWorksheet(
        title=SUMMARY_SHEET,
        cells=tuple(logical_rows),
        freeze_panes=None,
        auto_filter=None,
    )


def _validate_logical_invariants(
    model: LogicalWorkbook,
    result: ProcessingResult,
    summary: ProcessingSummary,
    base_columns: tuple[str, ...],
    error_column: str,
    output_path: Path,
) -> None:
    if tuple(sheet.title for sheet in model) != WORKSHEET_NAMES:
        _raise_data_error(output_path, "logical report has an invalid worksheet structure")

    expected_counts = (
        len(result.valid_records),
        len(result.invalid_records),
        len(result.duplicate_records),
    )
    for sheet, expected_count in zip(model[1:], expected_counts, strict=True):
        if len(sheet.cells) - 1 != expected_count:
            _raise_data_error(
                output_path,
                f"logical worksheet '{sheet.title}' has an invalid projection count",
            )

    expected_headers = (
        base_columns,
        (*base_columns, error_column),
        (*base_columns, error_column),
    )
    for sheet, columns in zip(model[1:], expected_headers, strict=True):
        headers = tuple(cell.value for cell in sheet.cells[0])
        generated_columns = frozenset(
            (*KNOWN_COLUMNS, *TRACEABILITY_COLUMNS, error_column)
        )
        independently_rendered_headers = tuple(
            f"'{column}"
            if column not in generated_columns and _is_formula_triggering_text(column)
            else column
            for column in columns
        )
        if headers != independently_rendered_headers or any(
            not isinstance(header, str) or header == "" for header in headers
        ):
            _raise_data_error(
                output_path,
                f"logical worksheet '{sheet.title}' has invalid headers",
            )
        if len(headers) != len(set(headers)):
            _raise_data_error(
                output_path,
                f"logical worksheet '{sheet.title}' has duplicate headers",
            )
        if not all(cell.bold for cell in sheet.cells[0]):
            _raise_data_error(
                output_path,
                f"logical worksheet '{sheet.title}' headers must be bold",
            )
        expected_filter = f"A1:{get_column_letter(len(columns))}{len(sheet.cells)}"
        if sheet.freeze_panes != "A2" or sheet.auto_filter != expected_filter:
            _raise_data_error(
                output_path,
                f"logical worksheet '{sheet.title}' has invalid navigation formatting",
            )

    summary_values = tuple(
        tuple(cell.value for cell in row)
        for row in model[0].cells[1:6]
    )
    expected_summary_values = (
        ("Total Records", summary.total_records),
        ("Valid Records", summary.valid_records),
        ("Invalid Records", summary.invalid_records),
        ("Duplicate Records", summary.duplicate_records),
        (
            "Total Paid Amount (USD)",
            _independent_expected_money_text(summary.total_paid_amount, output_path),
        ),
    )
    if summary_values != expected_summary_values:
        _raise_data_error(output_path, "logical summary differs from the supplied summary")

    amount_column_index = base_columns.index("amount")
    projections = (
        result.valid_records,
        result.invalid_records,
        result.duplicate_records,
    )
    for sheet, records in zip(model[1:], projections, strict=True):
        for row, processed_record in zip(sheet.cells[1:], records, strict=True):
            original_amount = processed_record.record.get("amount")
            if isinstance(original_amount, Decimal):
                expected_amount = _independent_expected_money_text(
                    original_amount,
                    output_path,
                )
                if _is_formula_triggering_text(expected_amount):
                    expected_amount = f"'{expected_amount}"
                if row[amount_column_index].value != expected_amount:
                    _raise_data_error(
                        output_path,
                        f"logical worksheet '{sheet.title}' has an invalid amount",
                    )
    if tuple(cell.value for cell in model[0].cells[0]) != ("Metric", "Value"):
        _raise_data_error(output_path, "logical summary has invalid headers")
    if tuple(cell.value for cell in model[0].cells[7]) != (
        "Invalid and duplicate record counts may overlap.",
        None,
    ):
        _raise_data_error(output_path, "logical summary has an invalid overlap note")

    for sheet in model:
        for row in sheet.cells:
            for cell in row:
                if cell.data_type == "f":
                    _raise_data_error(output_path, "logical report must not contain formulas")
                if (
                    cell.input_controlled
                    and isinstance(cell.value, str)
                    and _is_formula_triggering_text(cell.value)
                ):
                    _raise_data_error(
                        output_path,
                        "input-controlled formula text was not neutralized",
                    )


def _build_logical_report(
    result: ProcessingResult,
    summary: ProcessingSummary,
    output_path: Path,
) -> LogicalWorkbook:
    _validate_projection_integrity(result, output_path)
    base_columns, error_column = _record_columns(result, output_path)
    model: LogicalWorkbook = (
        _summary_model(summary, output_path),
        _record_worksheet_model(
            VALID_SHEET,
            result.valid_records,
            base_columns,
            error_column,
            output_path,
            include_errors=False,
        ),
        _record_worksheet_model(
            INVALID_SHEET,
            result.invalid_records,
            base_columns,
            error_column,
            output_path,
            include_errors=True,
        ),
        _record_worksheet_model(
            DUPLICATES_SHEET,
            result.duplicate_records,
            base_columns,
            error_column,
            output_path,
            include_errors=True,
        ),
    )
    _validate_logical_invariants(
        model,
        result,
        summary,
        base_columns,
        error_column,
        output_path,
    )
    return model


def _read_logical_workbook(workbook: Workbook) -> LogicalWorkbook:
    worksheets: list[_LogicalWorksheet] = []
    for worksheet in workbook.worksheets:
        rows: list[tuple[_LogicalCell, ...]] = []
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            logical_row: list[_LogicalCell] = []
            for cell in row:
                value = None if cell.value == "" else cell.value
                logical_row.append(
                    _LogicalCell(
                        value=value,
                        data_type=None if value is None else cell.data_type,
                        number_format=cell.number_format,
                        bold=bool(cell.font.bold),
                        input_controlled=False,
                    )
                )
            rows.append(tuple(logical_row))

        freeze_panes = worksheet.freeze_panes
        if isinstance(freeze_panes, Cell):
            freeze_panes = freeze_panes.coordinate
        worksheets.append(
            _LogicalWorksheet(
                title=worksheet.title,
                cells=tuple(rows),
                freeze_panes=freeze_panes,
                auto_filter=worksheet.auto_filter.ref,
            )
        )
    return tuple(worksheets)


def _render_logical_workbook(expected_workbook: LogicalWorkbook) -> Workbook:
    workbook = Workbook()
    workbook.iso_dates = True
    try:
        first_sheet = workbook.active
        for sheet_index, expected_sheet in enumerate(expected_workbook):
            worksheet = (
                first_sheet
                if sheet_index == 0
                else workbook.create_sheet(expected_sheet.title)
            )
            worksheet.title = expected_sheet.title
            for row_index, expected_row in enumerate(expected_sheet.cells, start=1):
                for column_index, expected_cell in enumerate(expected_row, start=1):
                    cell = worksheet.cell(row=row_index, column=column_index)
                    cell.value = expected_cell.value
                    cell.number_format = expected_cell.number_format
                    if expected_cell.bold:
                        cell.font = BOLD_FONT
            worksheet.freeze_panes = expected_sheet.freeze_panes
            worksheet.auto_filter.ref = expected_sheet.auto_filter
        return workbook
    except Exception as error:
        try:
            workbook.close()
        except OSError as cleanup_error:
            error.add_note(f"Workbook cleanup also failed: {cleanup_error}")
        raise


def _validate_saved_workbook_invariants(
    workbook: Workbook,
    expected_workbook: LogicalWorkbook,
    result: ProcessingResult,
    summary: ProcessingSummary,
    output_path: Path,
) -> None:
    if tuple(workbook.sheetnames) != WORKSHEET_NAMES:
        _raise_data_error(output_path, "saved report has an invalid worksheet structure")

    projection_counts = (
        len(result.valid_records),
        len(result.invalid_records),
        len(result.duplicate_records),
    )
    for sheet_name, expected_count in zip(
        WORKSHEET_NAMES[1:],
        projection_counts,
        strict=True,
    ):
        worksheet = workbook[sheet_name]
        if worksheet.max_row - 1 != expected_count:
            _raise_data_error(
                output_path,
                f"saved worksheet '{sheet_name}' has an invalid projection count",
            )

    for sheet_index, sheet_name in enumerate(WORKSHEET_NAMES[1:], start=1):
        worksheet = workbook[sheet_name]
        headers = tuple(cell.value for cell in worksheet[1])
        expected_headers = tuple(
            cell.value for cell in expected_workbook[sheet_index].cells[0]
        )
        if (
            headers != expected_headers
            or any(not isinstance(header, str) or header == "" for header in headers)
            or len(headers) != len(set(headers))
        ):
            _raise_data_error(
                output_path,
                f"saved worksheet '{sheet_name}' has invalid headers",
            )

    summary_sheet = workbook[SUMMARY_SHEET]
    expected_summary = (
        ("Metric", "Value"),
        ("Total Records", summary.total_records),
        ("Valid Records", summary.valid_records),
        ("Invalid Records", summary.invalid_records),
        ("Duplicate Records", summary.duplicate_records),
        (
            "Total Paid Amount (USD)",
            _independent_expected_money_text(
                summary.total_paid_amount,
                output_path,
            ),
        ),
    )
    actual_summary = tuple(
        tuple(summary_sheet.cell(row, column).value for column in (1, 2))
        for row in range(1, 7)
    )
    if actual_summary != expected_summary:
        _raise_data_error(output_path, "saved summary differs from the supplied summary")
    if summary_sheet["A8"].value != "Invalid and duplicate record counts may overlap.":
        _raise_data_error(output_path, "saved summary has an invalid overlap note")

    amount_column = KNOWN_COLUMNS.index("amount") + 1
    projections = (
        result.valid_records,
        result.invalid_records,
        result.duplicate_records,
    )
    for sheet_name, records in zip(
        WORKSHEET_NAMES[1:],
        projections,
        strict=True,
    ):
        worksheet = workbook[sheet_name]
        for row_index, processed_record in enumerate(records, start=2):
            original_amount = processed_record.record.get("amount")
            if isinstance(original_amount, Decimal):
                expected_amount = _independent_expected_money_text(
                    original_amount,
                    output_path,
                )
                if _is_formula_triggering_text(expected_amount):
                    expected_amount = f"'{expected_amount}"
                if worksheet.cell(row_index, amount_column).value != expected_amount:
                    _raise_data_error(
                        output_path,
                        f"saved worksheet '{sheet_name}' has an invalid amount",
                    )

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    _raise_data_error(output_path, "saved report must not contain formulas")


def _validate_temporary_report(
    temp_path: Path,
    output_path: Path,
    expected_workbook: LogicalWorkbook,
    result: ProcessingResult,
    summary: ProcessingSummary,
) -> None:
    validation_workbook: Workbook | None = None
    try:
        if not temp_path.is_file() or temp_path.stat().st_size == 0:
            error = OSError("temporary report is missing or empty")
            raise ReportExportError(
                output_path,
                "temporary report was not created correctly",
            ) from error
        validation_workbook = load_workbook(
            temp_path,
            read_only=False,
            data_only=False,
        )
        _validate_saved_workbook_invariants(
            validation_workbook,
            expected_workbook,
            result,
            summary,
            output_path,
        )
        actual_workbook = _read_logical_workbook(validation_workbook)
        if actual_workbook != expected_workbook:
            error = ValueError(
                "temporary report logical content differs from the expected workbook"
            )
            raise ReportExportError(
                output_path,
                "temporary report validation failed",
            ) from error
        validation_workbook.close()
        validation_workbook = None
    except ReportExportError:
        raise
    except (
        BadZipFile,
        InvalidFileException,
        OSError,
        ParseError,
        ValueError,
    ) as error:
        raise ReportExportError(
            output_path,
            f"temporary report could not be validated: {error}",
        ) from error
    finally:
        if validation_workbook is not None:
            active_error = sys.exc_info()[1]
            try:
                validation_workbook.close()
            except OSError as cleanup_error:
                if active_error is None:
                    raise ReportExportError(
                        output_path,
                        f"validation workbook could not be closed: {cleanup_error}",
                    ) from cleanup_error
                active_error.add_note(
                    f"Validation workbook cleanup also failed: {cleanup_error}"
                )


def _preserve_carriage_returns(temp_path: Path, output_path: Path) -> None:
    rewrite_path: Path | None = None
    try:
        try:
            descriptor, rewrite_name = tempfile.mkstemp(
                dir=temp_path.parent,
                prefix=f".{output_path.name}.",
                suffix=".rewrite.xlsx",
            )
            os.close(descriptor)
            rewrite_path = Path(rewrite_name)

            with ZipFile(temp_path, "r") as source_archive, ZipFile(
                rewrite_path,
                "w",
            ) as target_archive:
                for member in source_archive.infolist():
                    content = source_archive.read(member.filename)
                    if member.filename.endswith(".xml"):
                        content = content.replace(b"\r", b"&#13;")
                    target_archive.writestr(member, content)

            os.replace(rewrite_path, temp_path)
            rewrite_path = None
        except (BadZipFile, OSError, ValueError) as error:
            raise ReportExportError(
                output_path,
                f"temporary report text could not be preserved: {error}",
            ) from error
    finally:
        if rewrite_path is not None:
            try:
                rewrite_path.unlink(missing_ok=True)
            except OSError as cleanup_error:
                active_error = sys.exc_info()[1]
                if active_error is None:
                    raise ReportExportError(
                        output_path,
                        f"temporary rewrite could not be removed: {cleanup_error}",
                    ) from cleanup_error
                active_error.add_note(
                    f"Temporary rewrite cleanup also failed: {cleanup_error}"
                )


def export_report(
    result: ProcessingResult,
    summary: ProcessingSummary,
    output_path: Path,
) -> None:
    """Build and atomically publish the approved Excel report."""
    temp_path: Path | None = None
    workbook: Workbook | None = None
    workbook_closed = False

    try:
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            if output_path.is_dir():
                raise IsADirectoryError(f"output path is a directory: {output_path}")
            descriptor, temp_name = tempfile.mkstemp(
                dir=output_path.parent,
                prefix=f".{output_path.name}.",
                suffix=".tmp.xlsx",
            )
            temp_path = Path(temp_name)
            os.close(descriptor)
        except OSError as error:
            raise ReportExportError(
                output_path,
                f"output location could not be prepared: {error}",
            ) from error

        expected_workbook = _build_logical_report(result, summary, output_path)
        workbook = _render_logical_workbook(expected_workbook)
        try:
            workbook.save(temp_path)
        except OSError as error:
            raise ReportExportError(
                output_path,
                f"temporary report could not be saved: {error}",
            ) from error

        try:
            workbook.close()
            workbook_closed = True
            workbook = None
        except OSError as error:
            raise ReportExportError(
                output_path,
                f"workbook could not be closed: {error}",
            ) from error

        _preserve_carriage_returns(temp_path, output_path)
        _validate_temporary_report(
            temp_path,
            output_path,
            expected_workbook,
            result,
            summary,
        )
        try:
            os.replace(temp_path, output_path)
            temp_path = None
        except OSError as error:
            raise ReportExportError(
                output_path,
                f"report could not be published atomically: {error}",
            ) from error
    finally:
        active_error = sys.exc_info()[1]
        if workbook is not None and not workbook_closed:
            try:
                workbook.close()
            except OSError as cleanup_error:
                if active_error is None:
                    raise ReportExportError(
                        output_path,
                        f"workbook could not be closed: {cleanup_error}",
                    ) from cleanup_error
                active_error.add_note(
                    f"Workbook cleanup also failed: {cleanup_error}"
                )

        if temp_path is not None:
            try:
                temp_path.unlink(missing_ok=True)
            except OSError as cleanup_error:
                if active_error is None:
                    raise ReportExportError(
                        output_path,
                        f"temporary report could not be removed: {cleanup_error}",
                    ) from cleanup_error
                active_error.add_note(
                    f"Temporary report cleanup also failed: {cleanup_error}"
                )
