"""Render deterministic portfolio PNGs from the real demonstration report."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Final, Sequence

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


DEFAULT_WORKBOOK: Final = Path("data/demo/output/sales_report.xlsx")
DEFAULT_OUTPUT_DIR: Final = Path("portfolio/screenshots")
EXPECTED_WORKSHEETS: Final = (
    "Summary",
    "Valid Records",
    "Invalid Records",
    "Duplicates",
)

BACKGROUND: Final = "#F5F7FB"
SURFACE: Final = "#FFFFFF"
INK: Final = "#172033"
MUTED: Final = "#5D687A"
ACCENT: Final = "#2563EB"
ACCENT_DARK: Final = "#1746A2"
BORDER: Final = "#D9E0EA"
HEADER_FILL: Final = "#E8EEF8"
ROW_ALT: Final = "#F8FAFD"
INVALID: Final = "#C2413A"
DUPLICATE: Final = "#8B5CF6"
SUCCESS: Final = "#15803D"

FONT_REGULAR_PATH: Final = Path(
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
)
FONT_BOLD_PATH: Final = Path(
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
)


class PortfolioRenderError(Exception):
    """Raised when the workbook or rendered assets violate the visual contract."""


@dataclass(frozen=True)
class WorksheetData:
    """Immutable worksheet values used by the renderer."""

    title: str
    headers: tuple[str, ...]
    rows: tuple[tuple[object, ...], ...]


@dataclass(frozen=True)
class TableSpec:
    """Exact workbook projection and visual metadata for one table image."""

    filename: str
    worksheet: str
    title: str
    subtitle: str
    headers: tuple[str, ...]
    rows: tuple[tuple[object, ...], ...]
    column_weights: tuple[float, ...]
    accent: str
    dimensions: tuple[int, int] = (1800, 1000)


def _font(size: int, *, bold: bool = False) -> ImageFont.FreeTypeFont:
    path = FONT_BOLD_PATH if bold else FONT_REGULAR_PATH
    if not path.is_file():
        raise PortfolioRenderError(f"required font is unavailable: {path}")
    return ImageFont.truetype(str(path), size=size)


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _load_worksheet_data(workbook_path: Path) -> dict[str, WorksheetData]:
    if not workbook_path.is_file():
        raise PortfolioRenderError(f"workbook does not exist: {workbook_path}")

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        if tuple(workbook.sheetnames) != EXPECTED_WORKSHEETS:
            raise PortfolioRenderError(
                "workbook worksheets differ from the approved report structure"
            )

        worksheets: dict[str, WorksheetData] = {}
        for worksheet in workbook.worksheets:
            cell_rows = tuple(tuple(row) for row in worksheet.iter_rows())
            formulas = tuple(
                cell.coordinate
                for row in cell_rows
                for cell in row
                if cell.data_type == "f"
            )
            if formulas:
                raise PortfolioRenderError(
                    f"worksheet {worksheet.title!r} contains formulas: {formulas}"
                )
            if not cell_rows:
                raise PortfolioRenderError(
                    f"worksheet {worksheet.title!r} is unexpectedly empty"
                )
            headers = tuple(cell.value for cell in cell_rows[0])
            if any(not isinstance(header, str) or not header for header in headers):
                raise PortfolioRenderError(
                    f"worksheet {worksheet.title!r} contains invalid headers"
                )
            rows = tuple(
                tuple(cell.value for cell in row)
                for row in cell_rows[1:]
            )
            worksheets[worksheet.title] = WorksheetData(
                title=worksheet.title,
                headers=headers,
                rows=rows,
            )
        return worksheets
    finally:
        workbook.close()


def _project_columns(
    worksheet: WorksheetData,
    columns: Sequence[str],
) -> tuple[tuple[str, ...], tuple[tuple[object, ...], ...]]:
    missing = tuple(column for column in columns if column not in worksheet.headers)
    if missing:
        raise PortfolioRenderError(
            f"worksheet {worksheet.title!r} is missing visual columns: {missing}"
        )
    indexes = tuple(worksheet.headers.index(column) for column in columns)
    projected_rows = tuple(
        tuple(row[index] for index in indexes)
        for row in worksheet.rows
    )
    return tuple(columns), projected_rows


def _table_specs(worksheets: dict[str, WorksheetData]) -> tuple[TableSpec, ...]:
    valid_columns = (
        "order_id",
        "customer_name",
        "order_date",
        "amount",
        "status",
        "source_file",
        "source_sheet",
        "source_row",
    )
    review_columns = (
        "order_id",
        "customer_name",
        "amount",
        "status",
        "source_file",
        "source_sheet",
        "source_row",
        "validation_errors",
    )
    valid_headers, valid_rows = _project_columns(
        worksheets["Valid Records"], valid_columns
    )
    invalid_headers, invalid_rows = _project_columns(
        worksheets["Invalid Records"], review_columns
    )
    duplicate_headers, duplicate_rows = _project_columns(
        worksheets["Duplicates"], review_columns
    )
    return (
        TableSpec(
            filename="02-valid-records.png",
            worksheet="Valid Records",
            title="Validated Sales Records",
            subtitle=(
                f"{len(valid_rows)} unique records ready for reporting, "
                "with source traceability"
            ),
            headers=valid_headers,
            rows=valid_rows,
            column_weights=(1.15, 1.45, 1.15, 0.75, 0.85, 1.15, 1.2, 0.7),
            accent=SUCCESS,
        ),
        TableSpec(
            filename="03-invalid-records.png",
            worksheet="Invalid Records",
            title="Invalid Record Review",
            subtitle=(
                f"{len(invalid_rows)} records retained with explicit validation feedback"
            ),
            headers=invalid_headers,
            rows=invalid_rows,
            column_weights=(1.05, 1.25, 0.7, 0.8, 1.05, 1.0, 0.65, 3.0),
            accent=INVALID,
        ),
        TableSpec(
            filename="04-duplicates.png",
            worksheet="Duplicates",
            title="Duplicate Order Review",
            subtitle=(
                f"{len(duplicate_rows)} duplicate occurrences preserved without silent removal"
            ),
            headers=duplicate_headers,
            rows=duplicate_rows,
            column_weights=(1.05, 1.25, 0.7, 0.8, 1.05, 1.0, 0.65, 3.0),
            accent=DUPLICATE,
        ),
    )


def _wrap_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    font: ImageFont.FreeTypeFont,
    max_width: int,
) -> tuple[str, ...]:
    if not text:
        return ("",)
    words = text.replace("_", " ").split()
    if not words:
        return (text,)
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if draw.textlength(candidate, font=font) <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return tuple(lines)


def _draw_text_lines(
    draw: ImageDraw.ImageDraw,
    lines: Sequence[str],
    position: tuple[int, int],
    font: ImageFont.FreeTypeFont,
    fill: str,
    line_height: int,
) -> None:
    x, y = position
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_height


def _draw_brand_header(
    draw: ImageDraw.ImageDraw,
    title: str,
    subtitle: str,
    worksheet: str,
    accent: str,
    width: int,
) -> None:
    draw.rounded_rectangle((64, 54, 82, 164), radius=9, fill=accent)
    draw.text((110, 54), title, font=_font(42, bold=True), fill=INK)
    draw.text((110, 112), subtitle, font=_font(22), fill=MUTED)
    source = f"sales_report.xlsx  •  {worksheet}"
    source_width = int(draw.textlength(source, font=_font(18, bold=True)))
    draw.rounded_rectangle(
        (width - source_width - 116, 68, width - 64, 112),
        radius=20,
        fill=HEADER_FILL,
    )
    draw.text(
        (width - source_width - 90, 78),
        source,
        font=_font(18, bold=True),
        fill=ACCENT_DARK,
    )


def _render_summary(
    summary: WorksheetData,
    output_path: Path,
) -> tuple[tuple[str, object], ...]:
    if summary.headers != ("Metric", "Value"):
        raise PortfolioRenderError("Summary headers are not the approved headers")
    metrics = tuple(
        (row[0], row[1])
        for row in summary.rows
        if row[0] is not None and row[1] is not None
    )
    if len(metrics) != 5:
        raise PortfolioRenderError("Summary does not contain exactly five metrics")
    notes = tuple(row[0] for row in summary.rows if row[0] and row[1] is None)

    width, height = 1600, 900
    image = Image.new("RGB", (width, height), BACKGROUND)
    draw = ImageDraw.Draw(image)
    _draw_brand_header(
        draw,
        "Sales Processing Summary",
        "Automated CSV and XLSX validation, classification, and reporting",
        summary.title,
        ACCENT,
        width,
    )

    card_width = 448
    card_height = 210
    positions = (
        (64, 230),
        (576, 230),
        (1088, 230),
        (320, 486),
        (832, 486),
    )
    card_accents = (ACCENT, SUCCESS, INVALID, DUPLICATE, ACCENT_DARK)
    for (label, value), (x, y), card_accent in zip(
        metrics, positions, card_accents, strict=True
    ):
        draw.rounded_rectangle(
            (x, y, x + card_width, y + card_height),
            radius=24,
            fill=SURFACE,
            outline=BORDER,
            width=2,
        )
        draw.rounded_rectangle(
            (x + 24, y + 24, x + 34, y + card_height - 24),
            radius=5,
            fill=card_accent,
        )
        draw.text(
            (x + 58, y + 36),
            str(label),
            font=_font(22, bold=True),
            fill=MUTED,
        )
        draw.text(
            (x + 58, y + 92),
            _display_value(value),
            font=_font(54, bold=True),
            fill=INK,
        )

    if notes:
        draw.rounded_rectangle(
            (320, 748, 1280, 818),
            radius=18,
            fill=HEADER_FILL,
        )
        note = str(notes[0])
        note_width = draw.textlength(note, font=_font(19))
        draw.text(
            ((width - note_width) / 2, 772),
            note,
            font=_font(19),
            fill=ACCENT_DARK,
        )

    image.save(output_path, format="PNG", optimize=True)
    return metrics


def _column_widths(total_width: int, weights: Sequence[float]) -> tuple[int, ...]:
    weight_total = sum(weights)
    widths = [int(total_width * weight / weight_total) for weight in weights]
    widths[-1] += total_width - sum(widths)
    return tuple(widths)


def _render_table(spec: TableSpec, output_path: Path) -> None:
    width, height = spec.dimensions
    image = Image.new("RGB", spec.dimensions, BACKGROUND)
    draw = ImageDraw.Draw(image)
    _draw_brand_header(
        draw,
        spec.title,
        spec.subtitle,
        spec.worksheet,
        spec.accent,
        width,
    )

    table_x = 64
    table_y = 220
    table_width = width - 128
    header_height = 88
    cell_font = _font(18)
    header_font = _font(18, bold=True)
    line_height = 27
    padding_x = 16
    padding_y = 14
    column_widths = _column_widths(table_width, spec.column_weights)

    wrapped_headers = tuple(
        _wrap_text(draw, header, header_font, column_width - 2 * padding_x)
        for header, column_width in zip(
            spec.headers, column_widths, strict=True
        )
    )
    wrapped_rows = tuple(
        tuple(
            _wrap_text(
                draw,
                _display_value(value),
                cell_font,
                column_width - 2 * padding_x,
            )
            for value, column_width in zip(row, column_widths, strict=True)
        )
        for row in spec.rows
    )
    row_heights = tuple(
        max(66, max(len(lines) for lines in wrapped_row) * line_height + 2 * padding_y)
        for wrapped_row in wrapped_rows
    )
    table_height = header_height + sum(row_heights)
    if table_y + table_height > height - 54:
        raise PortfolioRenderError(
            f"table for {spec.worksheet!r} does not fit its canvas"
        )

    draw.rounded_rectangle(
        (table_x, table_y, table_x + table_width, table_y + table_height),
        radius=20,
        fill=SURFACE,
        outline=BORDER,
        width=2,
    )
    draw.rounded_rectangle(
        (table_x, table_y, table_x + table_width, table_y + header_height),
        radius=20,
        fill=HEADER_FILL,
    )
    draw.rectangle(
        (table_x, table_y + header_height - 20, table_x + table_width, table_y + header_height),
        fill=HEADER_FILL,
    )

    x = table_x
    for lines, column_width in zip(
        wrapped_headers, column_widths, strict=True
    ):
        _draw_text_lines(
            draw,
            lines,
            (x + padding_x, table_y + 18),
            header_font,
            ACCENT_DARK,
            24,
        )
        x += column_width

    y = table_y + header_height
    for row_index, (wrapped_row, row_height) in enumerate(
        zip(wrapped_rows, row_heights, strict=True)
    ):
        if row_index % 2:
            draw.rectangle(
                (table_x, y, table_x + table_width, y + row_height),
                fill=ROW_ALT,
            )
        x = table_x
        for lines, column_width in zip(
            wrapped_row, column_widths, strict=True
        ):
            _draw_text_lines(
                draw,
                lines,
                (x + padding_x, y + padding_y),
                cell_font,
                INK,
                line_height,
            )
            x += column_width
        y += row_height
        draw.line((table_x, y, table_x + table_width, y), fill=BORDER, width=1)

    x = table_x
    for column_width in column_widths[:-1]:
        x += column_width
        draw.line((x, table_y, x, table_y + table_height), fill=BORDER, width=1)

    image.save(output_path, format="PNG", optimize=True)


def _validate_semantics(
    worksheets: dict[str, WorksheetData],
    specs: Sequence[TableSpec],
    summary_metrics: Sequence[tuple[str, object]],
) -> None:
    expected_summary = tuple(
        (row[0], row[1])
        for row in worksheets["Summary"].rows
        if row[0] is not None and row[1] is not None
    )
    if tuple(summary_metrics) != expected_summary:
        raise PortfolioRenderError("rendered Summary metrics differ from the workbook")
    for spec in specs:
        headers, rows = _project_columns(worksheets[spec.worksheet], spec.headers)
        if headers != spec.headers or rows != spec.rows:
            raise PortfolioRenderError(
                f"rendered data differs from worksheet {spec.worksheet!r}"
            )


def _validate_pngs(
    output_dir: Path,
    specs: Sequence[TableSpec],
) -> None:
    expected = {
        "01-summary.png": (1600, 900),
        **{spec.filename: spec.dimensions for spec in specs},
    }
    for filename, dimensions in expected.items():
        path = output_dir / filename
        if not path.is_file() or path.stat().st_size == 0:
            raise PortfolioRenderError(f"rendered PNG is missing or empty: {path}")
        with Image.open(path) as image:
            image.verify()
        with Image.open(path) as image:
            if image.format != "PNG" or image.size != dimensions:
                raise PortfolioRenderError(
                    f"rendered PNG has invalid format or dimensions: {path}"
                )


def render_portfolio_screenshots(
    workbook_path: Path,
    output_dir: Path,
    *,
    validate_only: bool = False,
) -> None:
    """Render or validate the four approved worksheet-based portfolio assets."""
    worksheets = _load_worksheet_data(workbook_path)
    specs = _table_specs(worksheets)
    if validate_only:
        summary_metrics = tuple(
            (row[0], row[1])
            for row in worksheets["Summary"].rows
            if row[0] is not None and row[1] is not None
        )
    else:
        output_dir.mkdir(parents=True, exist_ok=True)
        summary_metrics = _render_summary(
            worksheets["Summary"], output_dir / "01-summary.png"
        )
        for spec in specs:
            _render_table(spec, output_dir / spec.filename)
    _validate_semantics(worksheets, specs, summary_metrics)
    _validate_pngs(output_dir, specs)


def _arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Render portfolio PNGs from the real demonstration report."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--validate-only", action="store_true")
    return parser.parse_args()


def main() -> None:
    """Run the deterministic portfolio renderer."""
    arguments = _arguments()
    try:
        render_portfolio_screenshots(
            arguments.workbook,
            arguments.output_dir,
            validate_only=arguments.validate_only,
        )
    except (OSError, ValueError, PortfolioRenderError) as error:
        raise SystemExit(f"Portfolio screenshot rendering failed: {error}") from error


if __name__ == "__main__":
    main()
